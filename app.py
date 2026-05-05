# v19
import os
from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_sqlalchemy import SQLAlchemy
import cloudinary
import cloudinary.uploader

# Cloudinary config — credentials set via Railway environment variables
cloudinary.config(
    cloud_name=os.environ.get('CLOUDINARY_CLOUD_NAME', ''),
    api_key=os.environ.get('CLOUDINARY_API_KEY', ''),
    api_secret=os.environ.get('CLOUDINARY_API_SECRET', ''),
    secure=True
)

def upload_to_cloudinary(file, folder='tahfeel-documents'):
    """Upload file to Cloudinary. Returns (url, public_id) or (None, None) on failure."""
    try:
        if not file or not file.filename: return None, None
        result = cloudinary.uploader.upload(
    file,
    folder=folder,
    resource_type='auto',
    use_filename=True,
    unique_filename=True,
    access_mode='public'
)
        return result.get('secure_url'), result.get('public_id')
    except Exception as e:
        print(f'Cloudinary upload error: {e}')
        return None, None
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, timezone
DUBAI_TZ = timezone(timedelta(hours=4))
def now_dubai():
    return datetime.now(DUBAI_TZ).replace(tzinfo=None)
from functools import wraps
import os

app = Flask(__name__)
app.secret_key = 'tahfeel2026secretkey'
# Session configuration for custom domain support
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_SECURE'] = True  # HTTPS only
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=7)
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///' + os.path.join(basedir, 'tahfeel.db')).replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    print(f"500 ERROR: {error}")
    import traceback
    traceback.print_exc()
    return "<h2>Something went wrong. Please <a href='/'>try again</a> or <a href='/logout'>logout and login</a>.</h2>", 500

@app.errorhandler(404)
def not_found(error):
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(100), unique=True, nullable=False)
    password = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='staff')
    active = db.Column(db.Boolean, default=True)
    phone = db.Column(db.String(20), nullable=True)

class Lead(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    company = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.String(200))
    source = db.Column(db.String(50))
    service = db.Column(db.String(100))
    representative = db.Column(db.String(100))
    lead_type = db.Column(db.String(20), default='New')
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime, default=lambda: now_dubai() + timedelta(days=1))
    remarks = db.Column(db.Text)
    status = db.Column(db.String(50), default='New')
    created_at = db.Column(db.DateTime, default=datetime.now)
    customer_story = db.Column(db.Text)
    phone2 = db.Column(db.String(20))
    campaign = db.Column(db.String(100))
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    updates = db.relationship('LeadUpdate', backref='lead', lazy=True, order_by='LeadUpdate.created_at.desc()')

class LeadUpdate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=False)
    stage = db.Column(db.String(50))
    remark = db.Column(db.Text, nullable=False)
    staff_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)
    followup_date = db.Column(db.DateTime)
    lost_reason = db.Column(db.String(100))
    future_potential = db.Column(db.String(20))

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')  # Low, Medium, High
    status = db.Column(db.String(20), default='Pending')   # Pending, In Progress, Done
    created_at = db.Column(db.DateTime, default=datetime.now)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    creator = db.relationship('User', foreign_keys=[created_by])
    lead = db.relationship('Lead', foreign_keys=[lead_id])

class Service(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Source(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Campaign(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class ServiceType(db.Model):
    __tablename__ = 'job_type'  # keep same DB table name
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    default_days = db.Column(db.Integer, default=1)

class Customer(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    company = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    phone2 = db.Column(db.String(20))
    email = db.Column(db.String(100))
    address = db.Column(db.String(200))
    source = db.Column(db.String(50))
    nationality = db.Column(db.String(50))
    date_of_birth = db.Column(db.Date, nullable=True)
    customer_type = db.Column(db.String(20), default='Individual')
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    lead_id = db.Column(db.Integer, db.ForeignKey('lead.id'), nullable=True)
    rep = db.relationship('User', foreign_keys=[assigned_to])
    lead = db.relationship('Lead', foreign_keys=[lead_id])
    jobs = db.relationship('Job', backref='customer', lazy=True, order_by='Job.created_at.desc()')

class Job(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=False)
    job_type = db.Column(db.String(100), nullable=False)
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'))
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')
    status = db.Column(db.String(50), default='Pending Finance Approval')
    internal_notes = db.Column(db.Text)
    service_note = db.Column(db.String(200))
    amount_invoiced = db.Column(db.Float, default=0)
    amount_received = db.Column(db.Float, default=0)
    num_persons = db.Column(db.Integer, default=1)
    finance_approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    finance_approved_at = db.Column(db.DateTime, nullable=True)
    finance_notes = db.Column(db.Text)
    # Completion fields
    final_remarks = db.Column(db.Text, nullable=True)
    future_work_notes = db.Column(db.Text, nullable=True)
    completed_at = db.Column(db.DateTime, nullable=True)
    # Partner commission fields
    partner_commission_expected = db.Column(db.Boolean, default=False)
    partner_name = db.Column(db.String(100))
    partner_amount = db.Column(db.Float)
    partner_due_date = db.Column(db.Date)
    partner_status = db.Column(db.String(20), default='Pending')  # Pending/Received/Written Off
    partner_received_date = db.Column(db.Date)
    revenue = db.Column(db.Float, default=0)  # Revenue counted when no partner OR when partner pays
    revenue_date = db.Column(db.Date)  # Date when revenue is counted (for cash-basis accounting)
    created_at = db.Column(db.DateTime, default=datetime.now)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    assignee = db.relationship('User', foreign_keys=[assigned_to])
    creator = db.relationship('User', foreign_keys=[created_by])
    finance_approver = db.relationship('User', foreign_keys=[finance_approved_by])
    updates = db.relationship('JobUpdate', backref='job', lazy=True, order_by='JobUpdate.created_at.desc()')
    subtasks = db.relationship('SubTask', backref='job', lazy=True, order_by='SubTask.created_at')
    partial_revenues = db.relationship('PartialRevenue', backref='job', lazy=True, order_by='PartialRevenue.revenue_date.desc()')

class PartialRevenue(db.Model):
    __tablename__ = 'partial_revenue'
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    amount = db.Column(db.Float, nullable=False)
    revenue_date = db.Column(db.Date, nullable=False)
    notes = db.Column(db.String(500))
    recorded_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)
    recorder = db.relationship('User', foreign_keys=[recorded_by])

class Partner(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.now)

class SubTask(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    title = db.Column(db.String(200), nullable=False)
    service_type = db.Column(db.String(100))
    assigned_to = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    due_date = db.Column(db.DateTime)
    priority = db.Column(db.String(20), default='Medium')
    status = db.Column(db.String(20), default='Pending')
    remarks = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    completed_at = db.Column(db.DateTime, nullable=True)
    assignee = db.relationship('User', foreign_keys=[assigned_to])

class DocType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)

class Document(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    doc_type = db.Column(db.String(100), nullable=False)
    belongs_to = db.Column(db.String(20), nullable=False)  # Company / Individual / Staff
    owner_name = db.Column(db.String(100), nullable=False)
    customer_id = db.Column(db.Integer, db.ForeignKey('customer.id'), nullable=True)
    expiry_date = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text)
    file_name = db.Column(db.String(255), nullable=True)
    file_url = db.Column(db.Text, nullable=True)
    cloudinary_public_id = db.Column(db.String(255), nullable=True)
    added_by = db.Column(db.String(100), nullable=True)
    uploaded_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    created_at = db.Column(db.DateTime, default=datetime.now)
    uploader = db.relationship('User', foreign_keys=[uploaded_by])
    customer = db.relationship('Customer', foreign_keys=[customer_id])


class ActivityLog(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    log_date = db.Column(db.Date, nullable=False)
    # Activity counts
    calls_existing = db.Column(db.Integer, default=0)
    calls_cold = db.Column(db.Integer, default=0)
    dm_instagram = db.Column(db.Integer, default=0)
    dm_facebook = db.Column(db.Integer, default=0)
    dm_linkedin = db.Column(db.Integer, default=0)
    posts_social = db.Column(db.Integer, default=0)
    videos_instagram = db.Column(db.Integer, default=0)
    linkedin_writing = db.Column(db.Integer, default=0)
    whatsapp_prospecting = db.Column(db.Integer, default=0)
    community_active = db.Column(db.Integer, default=0)
    google_reviews = db.Column(db.Integer, default=0)
    real_estate_relations = db.Column(db.Integer, default=0)
    content_marketing = db.Column(db.Integer, default=0)
    referral_building = db.Column(db.Integer, default=0)
    networking_activities = db.Column(db.Integer, default=0)
    networking_events = db.Column(db.Integer, default=0)
    off_day = db.Column(db.String(20), nullable=True)
    notes = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    user = db.relationship('User', foreign_keys=[user_id])


class ActivityType(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    field_key = db.Column(db.String(50), nullable=False, unique=True)
    label = db.Column(db.String(150), nullable=False)
    weekly_target = db.Column(db.Float, default=5)
    sort_order = db.Column(db.Integer, default=0)
    active = db.Column(db.Boolean, default=True)

class JobUpdate(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('job.id'), nullable=False)
    status = db.Column(db.String(50))
    status_note = db.Column(db.String(100))
    remark = db.Column(db.Text, nullable=False)
    staff_name = db.Column(db.String(100))
    created_at = db.Column(db.DateTime, default=datetime.now)

class MonthlyTarget(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    month = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    lead_target = db.Column(db.Integer, default=0)
    conversion_target = db.Column(db.Integer, default=0)
    amount_target = db.Column(db.Float, default=0)
    user = db.relationship('User', foreign_keys=[user_id])

class DeskNote(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    text = db.Column(db.Text, nullable=False)
    reminder_date = db.Column(db.Date, nullable=True)
    mention_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_done = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.now)
    user = db.relationship('User', foreign_keys=[user_id])
    mention_user = db.relationship('User', foreign_keys=[mention_user_id])

@app.context_processor
def inject_globals():
    result = {'birthdays_today': [], 'show_backup_reminder': False}
    try:
        if 'user_id' in session and session.get('role') == 'admin':
            last_backup = session.get('last_backup_date')
            if not last_backup:
                result['show_backup_reminder'] = True
            else:
                from datetime import datetime as _dt
                last = _dt.strptime(last_backup, '%Y-%m-%d')
                if (now_dubai() - last).days >= 14:
                    result['show_backup_reminder'] = True
    except:
        pass
    return result

@app.context_processor
def inject_birthdays():
    try:
        if 'user_id' in session:
            today = now_dubai()
            bdays = []
            try:
                result = db.session.execute(db.text(
                    "SELECT id, name, phone FROM customer WHERE date_of_birth IS NOT NULL AND EXTRACT(MONTH FROM date_of_birth)=:m AND EXTRACT(DAY FROM date_of_birth)=:d"
                ), {'m': today.month, 'd': today.day}).fetchall()
                bdays = [{'id': r[0], 'name': r[1], 'phone': r[2]} for r in result]
            except:
                try:
                    all_c = db.session.execute(db.text("SELECT id, name, phone, date_of_birth FROM customer WHERE date_of_birth IS NOT NULL")).fetchall()
                    bdays = [{'id':r[0],'name':r[1],'phone':r[2]} for r in all_c if r[3] and r[3].month==today.month and r[3].day==today.day]
                except:
                    pass
            return {'birthdays_today': bdays}
    except Exception as e:
        print(f'Birthday error: {e}')
    return {'birthdays_today': []}

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Admin access required')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def finance_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get('role') not in ['admin', 'finance']:
            flash('Finance access required')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

def apply_lead_filters(leads, args, now):
    date_filter = args.get('date')
    status_filter = args.get('status')
    staff_filter = args.get('staff')
    due_filter = args.get('due')  # NEW: Due Date filter
    
    if date_filter == 'today':
        leads = [l for l in leads if l.created_at and l.created_at.date() == now.date()]
    elif date_filter == 'week':
        week_start = now.date() - timedelta(days=now.weekday())
        week_end = week_start + timedelta(days=6)
        leads = [l for l in leads if l.created_at and week_start <= l.created_at.date() <= week_end]
    elif date_filter == 'month':
        leads = [l for l in leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
    elif date_filter == 'custom':
        from_date = args.get('from')
        to_date = args.get('to')
        if from_date:
            try:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                leads = [l for l in leads if l.created_at and l.created_at.date() >= from_dt]
            except: pass
        if to_date:
            try:
                to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                leads = [l for l in leads if l.created_at and l.created_at.date() <= to_dt]
            except: pass
    
    if status_filter:
        if status_filter == 'Overdue':
            leads = [l for l in leads if l.due_date < now and l.status not in ['Converted', 'Lost']]
        elif status_filter == 'Initiated':
            # Initiated = any action taken (not New, Converted, or Lost)
            leads = [l for l in leads if l.status not in ['New', 'Converted', 'Lost']]
        else:
            leads = [l for l in leads if l.status == status_filter]
    
    if staff_filter:
        try:
            sf = int(staff_filter)
            leads = [l for l in leads if l.assigned_to == sf]
        except: pass
    
    # NEW: Due Date filter
    if due_filter:
        if due_filter == 'overdue':
            leads = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost']]
        elif due_filter == 'today':
            leads = [l for l in leads if l.due_date and l.due_date.date() == now.date()]
        elif due_filter == 'tomorrow':
            tomorrow = now.date() + timedelta(days=1)
            leads = [l for l in leads if l.due_date and l.due_date.date() == tomorrow]
        elif due_filter == 'this_week':
            week_end = now.date() + timedelta(days=7)
            leads = [l for l in leads if l.due_date and now.date() <= l.due_date.date() <= week_end]
        elif due_filter == 'next_week':
            next_week_start = now.date() + timedelta(days=7)
            next_week_end = next_week_start + timedelta(days=7)
            leads = [l for l in leads if l.due_date and next_week_start <= l.due_date.date() <= next_week_end]
    
    return leads

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = User.query.filter_by(email=email, active=True).first()
        if user and check_password_hash(user.password, password):
            session.permanent = True  # Enable persistent session
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['role'] = user.role
            try:
                session['unread_mentions'] = DeskNote.query.filter_by(mention_user_id=user.id, is_done=False).count()
            except:
                session['unread_mentions'] = 0
            return redirect(url_for('dashboard'))
        flash('Invalid email or password')
        return render_template('login.html', email=email)
    return render_template('login.html', email='')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    now = now_dubai()
    role = session['role']

    # ── Finance dashboard ────────────────────────────────────────────────────
    if role == 'finance':
        date_filter = request.args.get('date', 'month')  # Default to current month
        
        try:
            all_jobs = Job.query.order_by(Job.created_at.desc()).all()
            
            # Filter jobs by date
            if date_filter == 'today':
                jobs = [j for j in all_jobs if j.created_at and j.created_at.date() == now.date()]
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())  # Monday
                week_end = week_start + timedelta(days=6)  # Sunday
                jobs = [j for j in all_jobs if j.created_at and week_start <= j.created_at.date() <= week_end]
            elif date_filter == 'month':
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
            elif date_filter == 'all':
                jobs = all_jobs
            else:
                # Default to month
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
            
            active_jobs = [j for j in jobs if j.status != 'Done']
            pending_approval = [j for j in jobs if j.status in ['Pending Finance Approval', 'Done']]
            pending_close = [j for j in jobs if j.status == 'Pending Finance Close']
            total_invoiced = sum((j.amount_invoiced or 0) for j in active_jobs)
            total_received = sum((j.amount_received or 0) for j in active_jobs)
            total_pending = total_invoiced - total_received
            completed_value = sum((j.amount_received or 0) for j in jobs if j.status == 'Done')
            
            # Revenue calculations (closed tasks + partial revenues from in-progress tasks)
            # Revenue filtered by revenue_date (when closed), not created_at
            if date_filter == 'today':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date == now.date()]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())
                week_end = week_start + timedelta(days=6)
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and week_start <= j.revenue_date <= week_end]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'month':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
                partial_revenue_jobs = all_jobs
            elif date_filter == 'all':
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
                partial_revenue_jobs = all_jobs
            else:
                # Default to month
                closed_jobs = [j for j in all_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission'] 
                              and j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
                partial_revenue_jobs = all_jobs
            
            total_revenue = sum((j.revenue or 0) for j in closed_jobs)
            
            # Add partial revenues from period (filter partial revenues by revenue_date too)
            partial_revenue_total = 0
            if date_filter == 'today':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if pr.revenue_date == now.date():
                            partial_revenue_total += pr.amount
            elif date_filter == 'week':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if week_start <= pr.revenue_date <= week_end:
                            partial_revenue_total += pr.amount
            elif date_filter == 'month':
                for j in partial_revenue_jobs:
                    for pr in j.partial_revenues:
                        if pr.revenue_date.year == now.year and pr.revenue_date.month == now.month:
                            partial_revenue_total += pr.amount
            else:  # all
                for j in partial_revenue_jobs:
                    partial_revenue_total += sum(pr.amount for pr in j.partial_revenues)
            
            total_revenue += partial_revenue_total
            
            # Customer Advances calculation (same as Admin dashboard)
            all_received = sum((j.amount_received or 0) for j in jobs)
            closed_received = sum((j.amount_received or 0) for j in closed_jobs)
            customer_advances = all_received - closed_received
            
            # Partner commission calculations
            partner_jobs = [j for j in all_jobs if j.partner_commission_expected and j.partner_status == 'Pending']
            total_partner_pending = sum((j.partner_amount or 0) for j in partner_jobs)
            
            # Monthly target
            targets = MonthlyTarget.query.filter_by(month=now.month, year=now.year).all()
            total_monthly_target = sum((t.amount_target or 0) for t in targets)
        except:
            active_jobs = pending_approval = pending_close = []
            total_invoiced = total_received = total_pending = completed_value = customer_advances = 0
            total_revenue = total_partner_pending = total_monthly_target = 0
        try:
            all_docs = Document.query.all()
            docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
            docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
            docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
            total_docs = len(all_docs)
        except:
            docs_30 = docs_60 = docs_90 = total_docs = 0
        all_active_jobs = Job.query.filter(Job.status.notin_(['Closed'])).all()
        tasks_active = len([j for j in all_active_jobs if j.status not in ['Pending Finance Approval','Pending Finance Close','Closed']])
        tasks_overdue = len([j for j in all_active_jobs if j.due_date and j.due_date < now and j.status not in ['Closed','Done']])
        tasks_processing = len([j for j in all_active_jobs if j.status == 'Processing'])
        tasks_pending_approval = len(pending_approval)
        return render_template('dashboard_finance.html',
                               now=now,
                               date_filter=date_filter,
                               docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                               all_jobs=active_jobs,
                               pending_approval=pending_approval,
                               pending_close=pending_close,
                               tasks_active=tasks_active,
                               tasks_overdue=tasks_overdue,
                               tasks_processing=tasks_processing,
                               tasks_pending_approval=tasks_pending_approval,
                               total_invoiced=total_invoiced,
                               total_received=total_received,
                               total_pending=total_pending,
                               completed_value=completed_value,
                               customer_advances=customer_advances,
                               birthdays_today=[],
                               total_revenue=total_revenue,
                               total_partner_pending=total_partner_pending,
                               total_monthly_target=total_monthly_target)

    # ── Admin dashboard ──────────────────────────────────────────────────────
    if role == 'admin':
        all_leads = Lead.query.order_by(Lead.due_date).all()
        date_filter = request.args.get('date', 'month')  # DEFAULT TO CURRENT MONTH
        from_date = request.args.get('from', '')
        to_date = request.args.get('to', '')

        leads = all_leads
        try:
            all_jobs = Job.query.order_by(Job.due_date).all()
            if date_filter == 'today':
                leads = [l for l in all_leads if l.created_at and l.created_at.date() == now.date()]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.date() == now.date()]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date == now.date()]
            elif date_filter == 'week':
                week_start = now.date() - timedelta(days=now.weekday())  # Monday
                week_end = week_start + timedelta(days=6)  # Sunday
                leads = [l for l in all_leads if l.created_at and week_start <= l.created_at.date() <= week_end]
                jobs = [j for j in all_jobs if j.created_at and week_start <= j.created_at.date() <= week_end]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and week_start <= j.revenue_date <= week_end]
            elif date_filter == 'month':
                leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
                # For revenue (cash-basis): use revenue_date instead of created_at
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
            elif date_filter == 'custom' and from_date and to_date:
                from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
                to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()
                leads = [l for l in all_leads if l.created_at and from_dt <= l.created_at.date() <= to_dt]
                jobs = [j for j in all_jobs if j.created_at and from_dt <= j.created_at.date() <= to_dt]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and from_dt <= j.revenue_date <= to_dt]
            elif date_filter == 'all':
                # Show all time only if explicitly selected
                leads = all_leads  # FIX: Include all leads
                jobs = all_jobs
                revenue_jobs = all_jobs
            else:
                # Default to current month
                leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
                jobs = [j for j in all_jobs if j.created_at and j.created_at.year == now.year and j.created_at.month == now.month]
                # For revenue (cash-basis): use revenue_date
                revenue_jobs = [j for j in all_jobs if j.revenue_date and j.revenue_date.year == now.year and j.revenue_date.month == now.month]
            active_jobs = [j for j in jobs if j.status not in ['Done', 'Closed', 'Closed - Pending Partner Commission']]
            done_jobs = [j for j in jobs if j.status == 'Done']
            closed_jobs = [j for j in jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
            # Revenue calculations use revenue_jobs (cash-basis)
            closed_revenue_jobs = [j for j in revenue_jobs if j.status in ['Closed', 'Closed - Pending Partner Commission']]
            
            # Finance totals: Count ALL jobs (active + done + closed) for consistency
            all_invoiced = sum((j.amount_invoiced or 0) for j in jobs)
            all_received = sum((j.amount_received or 0) for j in jobs)
            closed_received = sum((j.amount_received or 0) for j in closed_jobs)
            
            # Display values
            total_invoiced = all_invoiced
            total_received = all_received
            total_pending = all_invoiced - all_received
            completed_value = closed_received  # Money from CLOSED tasks (not Done)
            
            # Customer Advances = Money received but work not closed yet
            customer_advances = all_received - closed_received
            
            try:
                total_revenue = sum((j.revenue or 0) for j in closed_revenue_jobs)
                
                # Add partial revenues filtered by revenue_date (same period as revenue_jobs)
                partial_revenue_total = 0
                if date_filter == 'today':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if pr.revenue_date == now.date():
                                partial_revenue_total += pr.amount
                elif date_filter == 'week':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if week_start <= pr.revenue_date <= week_end:
                                partial_revenue_total += pr.amount
                elif date_filter == 'month':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if pr.revenue_date.year == now.year and pr.revenue_date.month == now.month:
                                partial_revenue_total += pr.amount
                elif date_filter == 'custom':
                    for j in all_jobs:
                        for pr in j.partial_revenues:
                            if from_dt <= pr.revenue_date <= to_dt:
                                partial_revenue_total += pr.amount
                else:  # all
                    for j in all_jobs:
                        partial_revenue_total += sum(pr.amount for pr in j.partial_revenues)
                
                total_revenue += partial_revenue_total
                
                # Partner commission pending
                partner_jobs = [j for j in all_jobs if j.partner_commission_expected and j.partner_status == 'Pending']
                total_partner_pending = sum((j.partner_amount or 0) for j in partner_jobs)
            except:
                total_revenue = 0
                total_partner_pending = 0
            overdue_jobs = [j for j in jobs if j.due_date and j.due_date < now and j.status not in ['Done', 'Pending Finance Approval']]
            pending_approval = [j for j in jobs if j.status == 'Pending Finance Approval']
            pending_close = [j for j in jobs if j.status == 'Pending Finance Close']
            recent_jobs = [j for j in all_jobs if j.status not in ['Done', 'Closed', 'Pending Finance Approval']][:10]
        except:
            jobs = all_jobs = active_jobs = done_jobs = closed_jobs = overdue_jobs = pending_approval = pending_close = recent_jobs = []
            total_invoiced = total_received = total_pending = completed_value = customer_advances = total_revenue = 0

        # Lead stats
        total = len(leads)
        overdue_leads = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost']]
        converted = [l for l in leads if l.status == 'Converted']
        lost = [l for l in leads if l.status == 'Lost']
        new_leads = [l for l in leads if l.status == 'New']
        # Initiated = any action taken (not New, Converted, or Lost)
        initiated = [l for l in leads if l.status not in ['New', 'Converted', 'Lost']]

        users = User.query.filter_by(active=True).filter(User.role != 'admin').all()
        # Workload — always this month
        wl_filter = 'month'
        wl_from = wl_to = ''
        all_leads_db = Lead.query.all()
        all_jobs_db = Job.query.all()
        def in_period(dt, f):
            if not dt: return False
            d = dt.date() if hasattr(dt,'date') else dt
            return d.month == now.month and d.year == now.year
        # Targets
        try:
            staff_targets = {t.user_id: t for t in MonthlyTarget.query.filter_by(month=now.month, year=now.year).all()}
            total_monthly_target = sum((t.amount_target or 0) for t in staff_targets.values())
        except:
            staff_targets = {}
            total_monthly_target = 0
        staff_stats = []
        for u in users:
            u_leads = [l for l in all_leads_db if l.assigned_to == u.id and in_period(l.created_at, wl_filter)]
            u_jobs_all = [j for j in all_jobs_db if j.assigned_to == u.id]
            u_jobs = [j for j in u_jobs_all if in_period(j.created_at, wl_filter)]
            u_closed = [j for j in u_jobs_all if j.status == 'Closed']
            # Sales value: credited to the customer's representative (assigned_to on customer)
            u_sales_jobs = [j for j in all_jobs_db if j.customer and j.customer.assigned_to == u.id and in_period(j.created_at, wl_filter)]
            u_sales_closed = [j for j in all_jobs_db if j.customer and j.customer.assigned_to == u.id and j.status == 'Closed']
            u_invoiced = sum((j.amount_invoiced or 0) for j in u_sales_jobs if j.status not in ['Pending Finance Approval'])
            u_closed_val = sum((j.amount_received or 0) for j in u_sales_closed)
            
            # Initiated = leads where staff took action (status NOT "New")
            u_initiated = len([l for l in u_leads if l.status != 'New'])
            
            # New leads = leads with status "New" (not yet contacted)
            u_new_leads = len([l for l in u_leads if l.status == 'New'])
            
            try:
                u_revenue = sum((j.revenue or 0) for j in u_sales_closed)
                # Add partial revenues from non-closed jobs for this staff
                for j in u_sales_jobs:
                    if j.status not in ['Closed', 'Closed - Pending Partner Commission']:
                        u_revenue += sum(pr.amount for pr in j.partial_revenues)
            except:
                u_revenue = 0
            t = staff_targets.get(u.id)
            amount_target = (t.amount_target or 0) if t else 0
            staff_stats.append({
                'name': u.name,
                'role': u.role,
                'leads': len(u_leads),
                'initiated': u_initiated,
                'new_leads': u_new_leads,
                'overdue_leads': len([l for l in u_leads if l.due_date and l.due_date < now and l.status not in ['Converted','Lost']]),
                'conversions': len([l for l in u_leads if l.status == 'Converted']),
                'lost': len([l for l in u_leads if l.status == 'Lost']),
                'active_jobs': len([j for j in u_jobs_all if j.status not in ['Done','Closed','Pending Finance Approval']]),
                'overdue_jobs': len([j for j in u_jobs_all if j.due_date and j.due_date < now and j.status not in ['Done','Closed','Pending Finance Approval']]),
                'invoiced': u_invoiced,
                'closed_val': u_closed_val,
                'revenue': u_revenue,
                'amount_target': amount_target,
                'leads_this_month': len(u_leads),
            })
        today_leads = [l for l in all_leads_db if l.created_at and l.created_at.date() == now.date()][:10]

        try:
            all_docs = Document.query.all()
            docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
            docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
            docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
            total_docs = len(all_docs)
        except:
            docs_30 = docs_60 = docs_90 = total_docs = 0
        # Birthdays today
        try:
            today = now.date()
            all_customers = Customer.query.filter(Customer.date_of_birth != None).all()
            birthdays_today = [c for c in all_customers if c.date_of_birth and c.date_of_birth.month == today.month and c.date_of_birth.day == today.day]
        except:
            birthdays_today = []
        return render_template('dashboard_admin.html',
                               leads=leads, today_leads=today_leads,
                               birthdays_today=birthdays_today,
                               wl_filter=wl_filter, wl_from=wl_from, wl_to=wl_to,
                               total=total, overdue_leads=overdue_leads,
                               converted=converted, lost=lost, new_leads=new_leads, initiated=initiated,
                               jobs=jobs, active_jobs=active_jobs,
                               overdue_jobs=overdue_jobs, done_jobs=done_jobs,
                               pending_approval=pending_approval,
                               pending_close=pending_close,
                               recent_jobs=recent_jobs,
                               total_invoiced=total_invoiced,
                               total_received=total_received,
                               total_pending=total_pending,
                               completed_value=completed_value,
                               customer_advances=customer_advances,
                               total_revenue=total_revenue,
                               total_partner_pending=total_partner_pending,
                               total_monthly_target=total_monthly_target,
                               staff_stats=staff_stats,
                               docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                               now=now, date_filter=date_filter,
                               from_date=from_date, to_date=to_date)

    # ── Staff dashboard ──────────────────────────────────────────────────────
    period = request.args.get('period', 'month')
    all_leads = Lead.query.filter_by(assigned_to=session['user_id']).order_by(Lead.due_date).all()
    if period == 'today':
        leads = [l for l in all_leads if l.created_at and l.created_at.date() == now.date()]
    elif period == 'week':
        week_start = now.date() - timedelta(days=now.weekday())  # Monday
        week_end = week_start + timedelta(days=6)  # Sunday
        leads = [l for l in all_leads if l.created_at and week_start <= l.created_at.date() <= week_end]
    elif period == 'month':
        leads = [l for l in all_leads if l.created_at and l.created_at.year == now.year and l.created_at.month == now.month]
    else:
        leads = all_leads
    overdue = [l for l in leads if l.due_date and l.due_date < now and l.status not in ['Converted', 'Lost']]
    converted = [l for l in leads if l.status == 'Converted']
    lost = [l for l in leads if l.status == 'Lost']
    new_leads = [l for l in leads if l.status == 'New']
    initiated = [l for l in leads if l.status not in ['New', 'Converted', 'Lost']]
    try:
        my_jobs = Job.query.filter_by(assigned_to=session['user_id']).filter(Job.status.notin_(['Done','Closed'])).order_by(Job.due_date).all()
        pending_approval_jobs = [j for j in my_jobs if j.status == 'Pending Finance Approval']
        overdue_jobs = [j for j in my_jobs if j.due_date and j.due_date < now and j.status != 'Pending Finance Approval']
        active_jobs = [j for j in my_jobs if j.status != 'Pending Finance Approval']
        total_invoiced = sum((j.amount_invoiced or 0) for j in active_jobs)
        total_received = sum((j.amount_received or 0) for j in active_jobs)
        total_pending = total_invoiced - total_received
        done_jobs = Job.query.filter_by(assigned_to=session['user_id'], status='Done').all()
        closed_jobs = Job.query.filter_by(assigned_to=session['user_id']).filter(Job.status.in_(['Closed', 'Closed - Pending Partner Commission'])).all()
        completed_value = sum((j.amount_received or 0) for j in done_jobs)
        try:
            total_revenue = sum((j.revenue or 0) for j in closed_jobs)
        except:
            total_revenue = 0
    except:
        my_jobs = []
        overdue_jobs = []
        total_invoiced = total_received = total_pending = completed_value = total_revenue = 0
    followups = LeadUpdate.query.filter(
        LeadUpdate.staff_name == session['user_name'],
        LeadUpdate.followup_date <= now + timedelta(days=1),
        LeadUpdate.followup_date >= now
    ).all()
    try:
        all_docs = Document.query.all()
        docs_30 = len([d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30])
        docs_60 = len([d for d in all_docs if d.expiry_date and 30 < (d.expiry_date - now).days <= 60])
        docs_90 = len([d for d in all_docs if d.expiry_date and 60 < (d.expiry_date - now).days <= 90])
        total_docs = len(all_docs)
    except:
        docs_30 = docs_60 = docs_90 = total_docs = 0
    try:
        today_date = now.date()
        all_customers_bday = Customer.query.filter(Customer.date_of_birth != None).all()
        birthdays_today = [c for c in all_customers_bday if c.date_of_birth and c.date_of_birth.month == today_date.month and c.date_of_birth.day == today_date.day]
    except:
        birthdays_today = []
    return render_template('dashboard_staff.html', leads=leads, overdue=overdue,
                           birthdays_today=birthdays_today,
                           converted=converted, lost=lost, new_leads=new_leads, initiated=initiated,
                           my_jobs=my_jobs, overdue_jobs=overdue_jobs,
                           total_invoiced=total_invoiced,
                           total_received=total_received,
                           total_pending=total_pending,
                           completed_value=completed_value,
                           total_revenue=total_revenue,
                           docs_30=docs_30, docs_60=docs_60, docs_90=docs_90, total_docs=total_docs,
                           pending_approval_jobs=pending_approval_jobs,
                           followups=followups, now=now, period=period)

@app.route('/leads')
@login_required
def all_leads():
    if session['role'] == 'finance':
        flash('Access denied')
        return redirect(url_for('dashboard'))
    now = now_dubai()
    leads = Lead.query.order_by(Lead.due_date).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    search = request.args.get('search', '').strip().lower()
    is_default = not any(request.args.get(k) for k in ['date', 'status', 'staff', 'search', 'from', 'to'])

    role = session.get('role')
    user_id = session.get('user_id')

    # For sales: default to their own leads unless staff filter explicitly set
    if role == 'sales' and not request.args.get('staff'):
        leads = [l for l in leads if l.assigned_to == user_id]

    if search:
        leads = [l for l in leads if
                 search in (l.name or '').lower() or
                 search in (l.phone or '').lower() or
                 search in (l.company or '').lower()]
    if is_default and request.args.get('date') != '':
        # Default: show this week's leads
        week_start = (now - timedelta(days=now.weekday())).date()
        leads = [l for l in leads if l.created_at and l.created_at.date() >= week_start]
    else:
        leads = apply_lead_filters(leads, request.args, now)

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 50
    total = len(leads)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = leads[(page - 1) * per_page: page * per_page]

    return render_template('all_leads.html', leads=paginated, now=now, users=users,
                           search=search, is_default=is_default,
                           page=page, total_pages=total_pages, total=total)

@app.route('/leads/export')
@login_required
def export_leads():
    if session['role'] not in ['admin', 'sales']:
        flash('Access denied')
        return redirect(url_for('all_leads'))
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import make_response
    import io
    now = now_dubai()
    leads = Lead.query.order_by(Lead.due_date).all()
    leads = apply_lead_filters(leads, request.args, now)
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    headers = ['Name', 'Company', 'Phone', 'Phone 2', 'Email', 'Address', 'Source',
               'Service', 'Lead Type', 'Assigned To', 'Due Date', 'Status', 'Remarks', 'Created']
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for lead in leads:
        ws.append([
            lead.name, lead.company or '', lead.phone or '', lead.phone2 or '',
            lead.email or '', lead.address or '', lead.source or '', lead.service or '',
            lead.lead_type or '', lead.assignee.name if lead.assignee else '',
            lead.due_date.strftime('%d %b %Y') if lead.due_date else '',
            lead.status or '', lead.remarks or '',
            lead.created_at.strftime('%d %b %Y') if lead.created_at else '',
        ])
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=tahfeel_leads_{now.strftime("%Y%m%d")}.xlsx'
    return response

@app.route('/leads/add', methods=['GET', 'POST'])
@login_required
def add_lead():
    now = now_dubai()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        due = request.form.get('due_date')
        lead_date = request.form.get('lead_date')
        created_dt = datetime.strptime(lead_date, '%Y-%m-%d') if lead_date else now_dubai()
        due_dt = datetime.strptime(due, '%Y-%m-%d') if due else created_dt + timedelta(days=1)
        lead = Lead(
            name=request.form['name'],
            company=request.form.get('company'),
            phone=request.form.get('phone'),
            phone2=request.form.get('phone2'),
            email=request.form.get('email'),
            address=request.form.get('address'),
            source=request.form.get('source'),
            service=request.form.get('service'),
            representative=session['user_name'],
            lead_type=request.form.get('lead_type', 'New'),
            assigned_to=int(request.form['assigned_to']) if request.form.get('assigned_to') else None,
            due_date=due_dt,
            remarks=request.form.get('remarks'),
            campaign=request.form.get('campaign') or None,
            created_at=created_dt
        )
        db.session.add(lead)
        db.session.commit()
        flash('Lead added successfully')
        return redirect(url_for('all_leads'))
    campaigns = Campaign.query.order_by(Campaign.name).all()
    tomorrow = (now_dubai() + timedelta(days=1)).strftime('%Y-%m-%d')
    return render_template('add_lead.html', users=users, services=services, sources=sources, campaigns=campaigns, now=now, tomorrow=tomorrow)

@app.route('/leads/<int:lead_id>', methods=['GET', 'POST'])
@login_required
def lead_detail(lead_id):
    now = now_dubai()
    lead = Lead.query.get_or_404(lead_id)
    if request.method == 'POST':
        stage = request.form['stage']
        remark = request.form['remark']
        followup = request.form.get('followup_date')
        followup_dt = datetime.strptime(followup, '%Y-%m-%d') if followup else None
        update = LeadUpdate(
            lead_id=lead.id, stage=stage, remark=remark,
            staff_name=session['user_name'], followup_date=followup_dt,
            lost_reason=request.form.get('lost_reason'),
            future_potential=request.form.get('future_potential')
        )
        lead.status = stage
        # Update lead's due_date with new followup date
        if followup_dt:
            lead.due_date = followup_dt
        if request.form.get('customer_story'):
            lead.customer_story = request.form.get('customer_story')
        db.session.add(update)
        db.session.commit()
        flash('Update saved')
        return redirect(url_for('lead_detail', lead_id=lead_id))
    return render_template('lead_detail.html', lead=lead, now=now)

@app.route('/leads/import', methods=['GET', 'POST'])
@login_required
def import_leads():
    if session.get('role') not in ['admin', 'sales']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or not file.filename.endswith('.xlsx'):
            flash('Please upload a valid .xlsx file')
            return redirect(url_for('import_leads'))
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file)
            ws = wb.active
            count = 0
            errors = []
            all_staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
            staff_map = {u.name.strip().lower(): u.id for u in all_staff}
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                name = row[0] if len(row) > 0 else None
                company = row[1] if len(row) > 1 else None
                phone = str(row[2]) if len(row) > 2 and row[2] else None
                email = row[3] if len(row) > 3 else None
                address = row[4] if len(row) > 4 else None
                source = row[5] if len(row) > 5 else None
                service = row[6] if len(row) > 6 else None
                lead_type = row[7] if len(row) > 7 else 'New'
                remarks = row[8] if len(row) > 8 else None
                assigned_name = str(row[9]).strip() if len(row) > 9 and row[9] else None
                lead_date_str = str(row[10]).strip() if len(row) > 10 and row[10] else None
                campaign = str(row[11]).strip() if len(row) > 11 and row[11] else None
                assigned_id = None
                if assigned_name:
                    assigned_id = staff_map.get(assigned_name.lower())
                    if not assigned_id:
                        errors.append(f'Row {i}: Staff "{assigned_name}" not found — lead imported unassigned')
                if not name:
                    errors.append(f'Row {i}: Name is required — skipped')
                    continue
                if not phone:
                    errors.append(f'Row {i}: Phone is required — skipped')
                    continue
                try:
                    if isinstance(lead_date_str, datetime):
                        created_dt = lead_date_str
                    elif lead_date_str:
                        created_dt = datetime.strptime(str(lead_date_str).split(' ')[0].split('T')[0], '%Y-%m-%d')
                    else:
                        created_dt = now_dubai()
                except:
                    created_dt = now_dubai()
                lead = Lead(
                    name=str(name), company=str(company) if company else None,
                    phone=str(phone), email=str(email) if email else None,
                    address=str(address) if address else None,
                    source=str(source) if source else None,
                    service=str(service) if service else None,
                    lead_type=str(lead_type) if lead_type else 'New',
                    remarks=str(remarks) if remarks else None,
                    representative=session['user_name'],
                    assigned_to=assigned_id,
                    campaign=campaign,
                    created_at=created_dt,
                    due_date=created_dt + timedelta(days=1)
                )
                db.session.add(lead)
                count += 1
            db.session.commit()
            if errors:
                flash(f'Imported {count} leads. Notes: ' + ' | '.join(errors))
            else:
                flash(f'Successfully imported {count} leads!')
            return redirect(url_for('dashboard'))
        except Exception as e:
            flash(f'Error reading file: {str(e)}')
            return redirect(url_for('import_leads'))
    return render_template('import_leads.html')

@app.route('/leads/template')
@login_required
def download_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
    from flask import make_response
    import io
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    staff = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    campaigns = Campaign.query.order_by(Campaign.name).all()
    headers = ['Name*', 'Company', 'Phone*', 'Email', 'Address',
               'Source', 'Service', 'Lead Type', 'Remarks', 'Assigned To', 'Lead Date', 'Campaign']
    ws.append(headers)
    ws.append(['John Smith', 'ABC Trading LLC', '+971501234567',
               'john@abc.ae', 'Dubai',
               sources[0].name if sources else 'WhatsApp',
               services[0].name if services else 'Trade License',
               'New', 'Interested in mainland license',
               staff[0].name if staff else '', '2026-04-16',
               campaigns[0].name if campaigns else ''])
    ws.append(['Sara Ahmed', '', '+971509876543', '', 'Sharjah',
               sources[1].name if len(sources) > 1 else '',
               services[1].name if len(services) > 1 else '',
               'New', '',
               staff[1].name if len(staff) > 1 else '', '2026-04-16', ''])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4
    ref = wb.create_sheet(title="Reference")
    ref['A1'] = 'Services'
    for i, s in enumerate(services, start=2):
        ref.cell(row=i, column=1, value=s.name)
    ref['B1'] = 'Sources'
    for i, s in enumerate(sources, start=2):
        ref.cell(row=i, column=2, value=s.name)
    ref['C1'] = 'Staff'
    for i, u in enumerate(staff, start=2):
        ref.cell(row=i, column=3, value=u.name)
    ref['D1'] = 'Lead Type'
    ref['D2'] = 'New'
    ref['D3'] = 'Old Follow-up'
    ref['E1'] = 'Campaigns'
    for i, c in enumerate(campaigns, start=2):
        ref.cell(row=i, column=5, value=c.name)
    ref.sheet_state = 'hidden'
    service_count = len(services) + 1
    source_count = len(sources) + 1
    staff_count = len(staff) + 1
    campaign_count = len(campaigns) + 1
    dv_source = DataValidation(type="list", formula1=f"Reference!$B$2:$B${source_count}", allow_blank=True, showDropDown=False)
    dv_source.sqref = "F2:F1000"
    ws.add_data_validation(dv_source)
    dv_service = DataValidation(type="list", formula1=f"Reference!$A$2:$A${service_count}", allow_blank=True, showDropDown=False)
    dv_service.sqref = "G2:G1000"
    ws.add_data_validation(dv_service)
    dv_type = DataValidation(type="list", formula1="Reference!$D$2:$D$3", allow_blank=True, showDropDown=False)
    dv_type.sqref = "H2:H1000"
    ws.add_data_validation(dv_type)
    dv_staff = DataValidation(type="list", formula1=f"Reference!$C$2:$C${staff_count}", allow_blank=True, showDropDown=False)
    dv_staff.sqref = "J2:J1000"
    ws.add_data_validation(dv_staff)
    if campaigns:
        dv_campaign = DataValidation(type="list", formula1=f"Reference!$E$2:$E${campaign_count}", allow_blank=True, showDropDown=False)
        dv_campaign.sqref = "L2:L1000"
        ws.add_data_validation(dv_campaign)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=tahfeel_leads_template.xlsx'
    return response

@app.route('/leads/<int:lead_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_lead(lead_id):
    now = now_dubai()
    lead = Lead.query.get_or_404(lead_id)
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        lead.name = request.form['name']
        lead.company = request.form.get('company')
        lead.phone = request.form.get('phone')
        lead.phone2 = request.form.get('phone2')
        lead.email = request.form.get('email')
        lead.address = request.form.get('address')
        lead.source = request.form.get('source')
        lead.service = request.form.get('service')
        lead.campaign = request.form.get('campaign') or None
        lead.lead_type = request.form.get('lead_type', 'New')
        lead.remarks = request.form.get('remarks')
        assigned = request.form.get('assigned_to')
        lead.assigned_to = int(assigned) if assigned else None
        due = request.form.get('due_date')
        if due:
            lead.due_date = datetime.strptime(due, '%Y-%m-%d')
        db.session.commit()
        flash('Lead updated successfully')
        return redirect(url_for('lead_detail', lead_id=lead_id))
    campaigns = Campaign.query.order_by(Campaign.name).all()
    return render_template('edit_lead.html', lead=lead, users=users, services=services, sources=sources, campaigns=campaigns, now=now)

@app.route('/leads/<int:lead_id>/delete')
@login_required
@admin_required
def delete_lead(lead_id):
    lead = Lead.query.get_or_404(lead_id)
    
    # Check if lead is converted (has associated customer)
    if lead.status == 'Converted':
        customer = Customer.query.filter_by(lead_id=lead_id).first()
        if customer:
            flash('Cannot delete converted leads. This lead has been converted to a customer and has associated records.', 'error')
            return redirect(url_for('lead_detail', lead_id=lead_id))
    
    LeadUpdate.query.filter_by(lead_id=lead_id).delete()
    db.session.delete(lead)
    db.session.commit()
    flash('Lead deleted successfully')
    ref = request.referrer
    if ref and '/leads' in ref:
        return redirect(ref)
    return redirect(url_for('all_leads'))

@app.route('/leads/bulk-delete', methods=['POST'])
@login_required
@admin_required
def bulk_delete_leads():
    ids = request.form.getlist('lead_ids')
    if not ids:
        flash('No leads selected', 'error')
        return redirect(url_for('all_leads'))
    count = 0
    skipped = 0
    for lead_id in ids:
        lead = Lead.query.get(int(lead_id))
        if lead:
            # Skip converted leads
            if lead.status == 'Converted':
                customer = Customer.query.filter_by(lead_id=lead.id).first()
                if customer:
                    skipped += 1
                    continue
            LeadUpdate.query.filter_by(lead_id=lead.id).delete()
            db.session.delete(lead)
            count += 1
    db.session.commit()
    if count > 0:
        flash(f'{count} lead(s) deleted successfully')
    if skipped > 0:
        flash(f'{skipped} converted lead(s) skipped (cannot delete converted leads)', 'warning')
    return redirect(url_for('all_leads'))

@app.route('/admin', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_panel():
    users = User.query.order_by(User.name).all()
    services = Service.query.order_by(Service.name).all()
    sources = Source.query.order_by(Source.name).all()
    campaigns = Campaign.query.order_by(Campaign.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    partners = Partner.query.order_by(Partner.name).all()
    return render_template('admin_panel.html', users=users, services=services,
                           sources=sources, campaigns=campaigns, job_types=job_types, doc_types=doc_types, partners=partners)

# TEMPORARILY DISABLED - WILL RE-ADD AFTER FIXING
# @app.route('/admin/import-data')
# @login_required
# @admin_required
# def import_data_page():
#     """Temporary page for importing historical Jan-March data"""
#     # Import tool for historical data
#     return render_template('import_data.html')

# @app.route('/admin/import-customers', methods=['POST'])
# @login_required
# @admin_required
# def import_customers():
#     """Import customers from Excel"""
#     from openpyxl import load_workbook
#     import io
#     
#     file = request.files.get('customers_file')
#     if not file:
#         flash('No file uploaded', 'error')
#         return redirect(url_for('import_data_page'))
#     
#     try:
#         wb = load_workbook(io.BytesIO(file.read()))
#         ws = wb.active
#         
#         # Get staff mapping
#         users = User.query.all()
#         staff_map = {u.name.lower(): u.id for u in users}
#         
#         imported = 0
#         skipped = 0
#         errors = []
#         
#         # Skip header row, start from row 2
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if not row[0]:  # Skip empty rows
#                 continue
#             
#             name = str(row[0]).strip() if row[0] else None
#             company = str(row[1]).strip() if row[1] else None
#             phone = str(row[2]).strip() if row[2] else None
#             email = str(row[3]).strip() if row[3] else None
#             assigned_to_name = str(row[4]).strip().lower() if row[4] else None
#             customer_type = str(row[5]).strip() if row[5] else 'Individual'
#             
#             if not name:
#                 skipped += 1
#                 continue
#             
#             # Check if customer already exists
#             existing = Customer.query.filter_by(phone=phone).first() if phone else None
#             if existing:
#                 skipped += 1
#                 continue
#             
#             # Map assigned_to
#             assigned_to_id = staff_map.get(assigned_to_name) if assigned_to_name else None
#             
#             customer = Customer(
#                 name=name,
#                 company=company,
#                 phone=phone,
#                 email=email,
#                 assigned_to=assigned_to_id,
#                 customer_type=customer_type,
#                 created_at=now_dubai()  # Will be backdated in tasks
#             )
#             db.session.add(customer)
#             imported += 1
#         
#         db.session.commit()
#         flash(f'✅ Imported {imported} customers. Skipped {skipped} (already exist or invalid).', 'success')
#     
#     except Exception as e:
#         db.session.rollback()
#         flash(f'Error: {str(e)}', 'error')
#     
#     return redirect(url_for('import_data_page'))

# @app.route('/admin/import-tasks', methods=['POST'])
# @login_required
# @admin_required  
# def import_tasks():
#     """Import tasks from Excel with historical dates"""
#     from openpyxl import load_workbook
#     import io
#     
#     file = request.files.get('tasks_file')
#     if not file:
#         flash('No file uploaded', 'error')
#         return redirect(url_for('import_data_page'))
#     
#     try:
#         wb = load_workbook(io.BytesIO(file.read()))
#         ws = wb.active
#         
#         # Get mappings
#         users = User.query.all()
#         staff_map = {u.name.lower(): u.id for u in users}
#         
#         customers = Customer.query.all()
#         customer_map = {c.name.lower(): c.id for c in customers}
#         
#         imported = 0
#         skipped = 0
#         
#         # Skip header row
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if not row[0]:
#                 continue
#             
#             customer_name = str(row[0]).strip().lower() if row[0] else None
#             job_type = str(row[1]).strip() if row[1] else None
#             status = str(row[2]).strip() if row[2] else 'Closed'
#             created_date = row[3] if row[3] else None
#             due_date = row[4] if row[4] else None
#             completed_date = row[5] if row[5] else None
#             assigned_to_name = str(row[6]).strip().lower() if row[6] else None
#             invoiced = float(row[7]) if row[7] else 0
#             received = float(row[8]) if row[8] else 0
#             priority = str(row[9]).strip() if row[9] else 'Medium'
#             remarks = str(row[10]).strip() if row[10] else None
#             
#             if not customer_name or not job_type:
#                 skipped += 1
#                 continue
#             
#             customer_id = customer_map.get(customer_name)
#             if not customer_id:
#                 skipped += 1
#                 continue
#             
#             assigned_to_id = staff_map.get(assigned_to_name) if assigned_to_name else None
#             
#             # Parse dates
#             try:
#                 if isinstance(created_date, str):
#                     created_dt = datetime.strptime(created_date, '%d/%m/%Y')
#                 else:
#                     created_dt = created_date
#             except:
#                 created_dt = now_dubai()
#             
#             try:
#                 if isinstance(due_date, str):
#                     due_dt = datetime.strptime(due_date, '%d/%m/%Y')
#                 else:
#                     due_dt = due_date
#             except:
#                 due_dt = None
#             
#             try:
#                 if isinstance(completed_date, str):
#                     completed_dt = datetime.strptime(completed_date, '%d/%m/%Y')
#                 else:
#                     completed_dt = completed_date
#             except:
#                 completed_dt = None
#             
#             # Calculate revenue (cash-basis)
#             revenue = received if status == 'Closed' else 0
#             
#             job = Job(
#                 customer_id=customer_id,
#                 job_type=job_type,
#                 status=status,
#                 assigned_to=assigned_to_id,
#                 created_by=assigned_to_id,
#                 priority=priority,
#                 created_at=created_dt,
#                 due_date=due_dt,
#                 completed_at=completed_dt,
#                 amount_invoiced=invoiced,
#                 amount_received=received,
#                 revenue=revenue,
#                 remarks=remarks
#             )
#             db.session.add(job)
#             imported += 1
#         
#         db.session.commit()
#         flash(f'✅ Imported {imported} tasks. Skipped {skipped} (missing customer or invalid data).', 'success')
#     
#     except Exception as e:
#         db.session.rollback()
#         flash(f'Error: {str(e)}', 'error')
#     
#     return redirect(url_for('import_data_page'))



@app.route('/admin/fix-cloudinary-access')
@login_required
@admin_required
def fix_cloudinary_access():
    """Fix Cloudinary document access - set all to public"""
    try:
        import cloudinary.uploader
        import cloudinary.api
        import re
        
        # Get all documents with Cloudinary URLs
        documents = Document.query.filter(Document.file_url.like('%cloudinary.com%')).all()
        fixed = 0
        errors = []
        
        for doc in documents:
            try:
                # Extract public_id from URL
                public_id = doc.cloudinary_public_id
                
                if not public_id and doc.file_url:
                    # Extract from URL: https://res.cloudinary.com/.../tahfeel-documents/document_file_nbxoqf.pdf
                    match = re.search(r'tahfeel-documents/[^.?]+', doc.file_url)
                    if match:
                        public_id = match.group(0)
                        doc.cloudinary_public_id = public_id
                        db.session.add(doc)
                
                if public_id:
                    try:
                        # Use Cloudinary API to update access control to public
                        # Try as image first (most common)
                        try:
                            result = cloudinary.uploader.explicit(
                                public_id,
                                type='upload',
                                resource_type='image',
                                access_control=[{"access_type": "anonymous"}]
                            )
                            fixed += 1
                            print(f"✓ Fixed Doc {doc.id}: {public_id}")
                        except Exception as e1:
                            # If image fails, try as raw (PDF)
                            if 'not found' in str(e1).lower():
                                result = cloudinary.uploader.explicit(
                                    public_id,
                                    type='upload',
                                    resource_type='raw',
                                    access_control=[{"access_type": "anonymous"}]
                                )
                                fixed += 1
                                print(f"✓ Fixed Doc {doc.id} as raw: {public_id}")
                            else:
                                raise e1
                    except Exception as e:
                        error_msg = f"Doc {doc.id} - {doc.doc_type or 'Unknown'}: {str(e)}"
                        errors.append(error_msg)
                        print(f"✗ Error: {error_msg}")
            
            except Exception as e:
                errors.append(f"Doc {doc.id}: {str(e)}")
        
        db.session.commit()
        
        if errors:
            error_summary = '<br>'.join(errors[:3])
            if len(errors) > 3:
                error_summary += f'<br>...and {len(errors)-3} more'
            flash(f'✓ Fixed {fixed} documents | ✗ {len(errors)} errors:<br>{error_summary}', 'warning')
        elif fixed > 0:
            flash(f'✓ Successfully set {fixed} documents to public access!', 'success')
        else:
            flash('No documents found to fix.', 'info')
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
        print(f"Critical Error: {e}")
        import traceback
        traceback.print_exc()
    
    return redirect(url_for('admin_panel'))


@app.route('/admin/staff/add', methods=['POST'])
@login_required
@admin_required
def admin_add_staff():
    email = request.form['email']
    existing = User.query.filter_by(email=email).first()
    if existing:
        flash('This email already exists')
        return redirect(url_for('admin_panel'))
    try:
        user = User(
            name=request.form['name'],
            email=email,
            password=generate_password_hash(request.form['password']),
            role=request.form.get('role', 'staff'),
            phone=request.form.get('phone', '').strip() or None
        )
        db.session.add(user)
        db.session.commit()
        flash('Staff member added successfully')
    except Exception as e:
        db.session.rollback()
        flash('Error — ' + str(e))
    return redirect(url_for('admin_panel'))

@app.route('/admin/staff/<int:user_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_staff(user_id):
    user = User.query.get_or_404(user_id)
    name = request.form.get('name', '').strip()
    email = request.form.get('email', '').strip()
    new_password = request.form.get('password', '').strip()
    new_role = request.form.get('role', '').strip()
    if name:
        user.name = name
    if email:
        existing = User.query.filter_by(email=email).first()
        if existing and existing.id != user_id:
            flash('That email is already in use')
            return redirect(url_for('admin_panel'))
        user.email = email
    if new_role in ['staff', 'sales', 'operations', 'admin', 'finance']:
        user.role = new_role
    if new_password:
        user.password = generate_password_hash(new_password)
    user.phone = request.form.get('phone', '').strip() or None
    db.session.commit()
    flash('Staff member updated successfully')
    return redirect(url_for('admin_panel'))

@app.route('/admin/service/add', methods=['POST'])
@login_required
@admin_required
def admin_add_service():
    name = request.form.get('name', '').strip()
    if name:
        existing = Service.query.filter_by(name=name).first()
        if existing:
            flash('Service already exists')
        else:
            db.session.add(Service(name=name))
            db.session.commit()
            flash(f'Service "{name}" added')
    return redirect(url_for('admin_panel'))

@app.route('/admin/service/<int:service_id>/delete')
@login_required
@admin_required
def admin_delete_service(service_id):
    service = Service.query.get_or_404(service_id)
    db.session.delete(service)
    db.session.commit()
    flash(f'Service "{service.name}" removed')
    return redirect(url_for('admin_panel'))

@app.route('/admin/source/add', methods=['POST'])
@login_required
@admin_required
def admin_add_source():
    name = request.form.get('name', '').strip()
    if name:
        existing = Source.query.filter_by(name=name).first()
        if existing:
            flash('Source already exists')
        else:
            db.session.add(Source(name=name))
            db.session.commit()
            flash(f'Source "{name}" added')
    return redirect(url_for('admin_panel'))

@app.route('/admin/source/<int:source_id>/delete')
@login_required
@admin_required
def admin_delete_source(source_id):
    source = Source.query.get_or_404(source_id)
    db.session.delete(source)
    db.session.commit()
    flash(f'Source "{source.name}" removed')
    return redirect(url_for('admin_panel'))

@app.route('/users/<int:user_id>/toggle')
@login_required
@admin_required
def toggle_user(user_id):
    user = User.query.get_or_404(user_id)
    user.active = not user.active
    db.session.commit()
    flash(f'{"Activated" if user.active else "Deactivated"} {user.name}')
    return redirect(url_for('admin_panel'))

@app.route('/admin/staff/<int:user_id>/toggle')
@login_required
@admin_required
def admin_toggle_staff(user_id):
    user = User.query.get_or_404(user_id)
    user.active = not user.active
    db.session.commit()
    flash(f'{"Activated" if user.active else "Deactivated"} {user.name}')
    return redirect(url_for('admin_panel'))

# ── Customers ────────────────────────────────────────────────────────────────

@app.route('/customers')
@login_required
def customers():
    now = now_dubai()
    search = request.args.get('search', '').strip().lower()
    birthday_filter = request.args.get('birthday', '')
    birthdays_today = []
    try:
        # Ensure columns exist first
        with db.engine.connect() as conn:
            for col, typ in [('phone2','VARCHAR(20)'),('assigned_to','INTEGER'),('date_of_birth','DATE')]:
                try:
                    conn.execute(db.text(f'ALTER TABLE customer ADD COLUMN IF NOT EXISTS {col} {typ}'))
                    conn.commit()
                except: pass
        customer_list = Customer.query.order_by(Customer.created_at.desc()).all()
        if search:
            customer_list = [c for c in customer_list if
                search in (c.name or '').lower() or
                search in (c.company or '').lower() or
                search in (c.phone or '').lower()]
        if birthday_filter == 'today':
            customer_list = [c for c in customer_list if c.date_of_birth and
                             c.date_of_birth.month == now.month and
                             c.date_of_birth.day == now.day]
        try:
            bday_list = Customer.query.filter(Customer.date_of_birth != None).all()
            birthdays_today = [c for c in bday_list if c.date_of_birth and
                               c.date_of_birth.month == now.month and c.date_of_birth.day == now.day]
        except:
            birthdays_today = []
    except Exception as e:
        customer_list = []
        flash(f'Error loading customers: {e}')
    page = int(request.args.get('page', 1))
    per_page = 25
    total = len(customer_list)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = customer_list[(page-1)*per_page : page*per_page]
    return render_template('customers.html', customers=paginated, page=page, total_pages=total_pages,
                           total=total, search=request.args.get('search',''),
                           birthdays_today=birthdays_today, now=now, birthday_filter=birthday_filter)

@app.route('/customers/add', methods=['GET', 'POST'])
@login_required
def add_customer():
    converted_leads = Lead.query.filter_by(status='Converted').order_by(Lead.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        # Validate required fields
        if not request.form.get('lead_id'):
            if not request.form.get('source'):
                flash('Source is required', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template('add_customer.html', users=users, sources=sources, converted_leads=converted_leads)
            if not request.form.get('assigned_to'):
                flash('Primary Representative is required', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template('add_customer.html', users=users, sources=sources, converted_leads=converted_leads)
        
        # Check for duplicate phone number
        phone_to_check = None
        lead_id = request.form.get('lead_id') or None
        if lead_id:
            lead = Lead.query.get(int(lead_id))
            phone_to_check = lead.phone
        else:
            phone_to_check = request.form.get('phone', '').strip()
        
        if phone_to_check:
            existing_customer = Customer.query.filter_by(phone=phone_to_check).first()
            if existing_customer:
                flash(f'⚠️ Duplicate customer! Phone number {phone_to_check} already exists for customer: {existing_customer.name}', 'error')
                users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
                return render_template('add_customer.html', users=users, sources=sources, converted_leads=converted_leads)
        
        if lead_id:
            customer = Customer(
                name=lead.name, company=lead.company, phone=lead.phone,
                phone2=getattr(lead, 'phone2', None),
                email=lead.email, address=lead.address, source=lead.source,
                notes=request.form.get('notes'), lead_id=int(lead_id)
            )
        else:
            customer = Customer(
                name=request.form.get('name', '').strip(),
                company=request.form.get('company', '').strip() or None,
                phone=request.form.get('phone', '').strip(),
                phone2=request.form.get('phone2', '').strip() or None,
                email=request.form.get('email', '').strip() or None,
                address=request.form.get('address', '').strip() or None,
                source=request.form.get('source', '').strip() or None,
                nationality=request.form.get('nationality', '').strip() or None,
                customer_type=request.form.get('customer_type', 'Individual'),
                assigned_to=int(request.form.get('assigned_to')) if request.form.get('assigned_to') else None,
                notes=request.form.get('notes', '').strip() or None,
                date_of_birth=datetime.strptime(request.form.get('date_of_birth'), '%Y-%m-%d').date() if request.form.get('date_of_birth') else None
            )
        db.session.add(customer)
        db.session.flush()  # get customer.id before commit

        # Save inline documents
        doc_types_inline = request.form.getlist('doc_type[]')
        doc_owners = request.form.getlist('doc_owner[]')
        doc_expiries = request.form.getlist('doc_expiry[]')
        doc_notes_list = request.form.getlist('doc_notes[]')

        for i, dt in enumerate(doc_types_inline):
            if not dt: continue
            expiry = None
            try:
                if i < len(doc_expiries) and doc_expiries[i]:
                    expiry = datetime.strptime(doc_expiries[i], '%Y-%m-%d')
            except: pass
            file_url, file_name, public_id = None, None, None
            doc_file = request.files.get(f'doc_file_{i+1}')
            if doc_file and doc_file.filename:
                file_url, public_id = upload_to_cloudinary(doc_file)
                file_name = doc_file.filename
            doc = Document(
                customer_id=customer.id,
                doc_type=dt,
                owner_name=doc_owners[i] if i < len(doc_owners) else customer.name,
                belongs_to='Individual',
                expiry_date=expiry,
                notes=doc_notes_list[i] if i < len(doc_notes_list) else None,
                added_by=session['user_name'],
                file_url=file_url,
                file_name=file_name,
                cloudinary_public_id=public_id,
                uploaded_by=session.get('user_id')
            )
            db.session.add(doc)

        db.session.commit()
        flash('Customer added successfully')
        return redirect(url_for('customer_detail', customer_id=customer.id))
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    return render_template('add_customer.html', converted_leads=converted_leads, sources=sources, users=users, doc_types=doc_types)

@app.route('/customers/<int:customer_id>')
@login_required
def customer_detail(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    now = now_dubai()
    jobs = Job.query.filter_by(customer_id=customer_id).order_by(Job.created_at.desc()).all()
    docs = Document.query.filter_by(customer_id=customer_id).order_by(Document.expiry_date).all()
    total_invoiced = sum(j.amount_invoiced or 0 for j in jobs)
    total_received = sum(j.amount_received or 0 for j in jobs)
    return render_template('customer_detail.html', customer=customer, jobs=jobs,
                           documents=docs, now=now,
                           total_invoiced=total_invoiced, total_received=total_received)


@app.route('/customers/<int:customer_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    sources = Source.query.order_by(Source.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    if request.method == 'POST':
        customer.name = request.form.get('name', '').strip() or customer.name
        customer.company = request.form.get('company', '').strip()
        customer.phone = request.form.get('phone', '').strip()
        customer.phone2 = request.form.get('phone2', '').strip() or None
        customer.email = request.form.get('email', '').strip()
        customer.address = request.form.get('address', '').strip()
        customer.source = request.form.get('source', '').strip()
        customer.nationality = request.form.get('nationality', '').strip() or None
        dob_str = request.form.get('date_of_birth', '').strip()
        customer.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date() if dob_str else None
        customer.customer_type = request.form.get('customer_type', 'Individual')
        try:
            customer.assigned_to = int(request.form.get('assigned_to')) if request.form.get('assigned_to') else None
        except:
            pass
        customer.notes = request.form.get('notes', '').strip()
        # Save any inline documents added
        doc_types_inline = request.form.getlist('doc_type[]')
        doc_owners = request.form.getlist('doc_owner[]')
        doc_expiries = request.form.getlist('doc_expiry[]')
        doc_notes_list = request.form.getlist('doc_notes[]')
        for i, dt in enumerate(doc_types_inline):
            if not dt: continue
            expiry = None
            try:
                if i < len(doc_expiries) and doc_expiries[i]:
                    expiry = datetime.strptime(doc_expiries[i], '%Y-%m-%d')
            except: pass
            # Handle file upload for this doc
            doc_file_key = f'doc_file_{i+1}'
            file_name = None
            file_url = None
            public_id = None
            if doc_file_key in request.files:
                f = request.files[doc_file_key]
                if f and f.filename:
                    file_name = f.filename
                    file_url, public_id = upload_to_cloudinary(f)
            doc = Document(
                customer_id=customer_id,
                doc_type=dt,
                owner_name=doc_owners[i] if i < len(doc_owners) and doc_owners[i] else customer.name,
                belongs_to='Individual',
                expiry_date=expiry,
                notes=doc_notes_list[i] if i < len(doc_notes_list) else None,
                file_name=file_name,
                file_url=file_url,
                cloudinary_public_id=public_id,
                added_by=session['user_name']
            )
            db.session.add(doc)
        db.session.commit()
        flash('Customer updated successfully')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    doc_types = DocType.query.order_by(DocType.name).all()
    existing_docs = Document.query.filter_by(customer_id=customer_id).order_by(Document.expiry_date).all()
    now = now_dubai()
    return render_template('edit_customer.html', customer=customer, sources=sources, users=users, doc_types=doc_types, existing_docs=existing_docs, now=now)

@app.route('/customers/<int:customer_id>/delete')
@login_required
@admin_required
def delete_customer(customer_id):
    customer = Customer.query.get_or_404(customer_id)
    # Only delete if no jobs linked
    if customer.jobs:
        flash('Cannot delete customer with existing tasks. Remove tasks first.')
        return redirect(url_for('customer_detail', customer_id=customer_id))
    Document.query.filter_by(customer_id=customer_id).delete()
    db.session.delete(customer)
    db.session.commit()
    flash('Customer deleted')
    return redirect(url_for('customers'))


@app.route('/customers/export')
@login_required
def export_customers():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Customers'
    headers = ['Name','Company','Phone','Phone 2','Email','Address','Nationality','Date of Birth','Customer Type','Source','Primary Representative','Notes']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    customers = Customer.query.order_by(Customer.created_at.desc()).all()
    for r, c in enumerate(customers, 2):
        # Get assigned representative name
        rep_name = ''
        if c.assigned_to:
            rep_user = User.query.get(c.assigned_to)
            if rep_user:
                rep_name = rep_user.name
        ws.append([
            c.name,
            c.company or '',
            c.phone or '',
            c.phone2 or '',
            c.email or '',
            c.address or '',
            c.nationality or '',
            c.date_of_birth.strftime('%d/%m/%Y') if c.date_of_birth else '',
            c.customer_type or 'Individual',
            c.source or '',
            rep_name,
            c.notes or ''
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='tahfeel_customers.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/customers/template')
@login_required
def customer_import_template():
    import io
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = 'Customers'
    from openpyxl.worksheet.datavalidation import DataValidation
    headers = ['Name *', 'Phone *', 'Company', 'Phone 2', 'Email', 'Address', 'Nationality', 'Date of Birth', 'Type', 'Source', 'Primary Representative *', 'Notes']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1A3B8B')
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 18)
    # Sample rows
    samples = [
        ['Ahmed Al Mansoori', '+971501234567', 'Al Mansoori Trading LLC', '+971551234567', 'ahmed@example.com', 'Dubai, UAE', 'Emirati', '15/01/1985', 'Company', 'Referral', 'Aslam', 'VIP client'],
        ['Priya Sharma', '+971507654321', '', '', 'priya@gmail.com', 'Sharjah, UAE', 'Indian', '20/05/1990', 'Individual', 'WhatsApp', 'Anfal', ''],
        ['XYZ Investments', '+971509876543', 'XYZ Investments LLC', '', '', 'Abu Dhabi, UAE', 'British', '', 'Company', 'Website', 'Lukman', 'Golden visa interested'],
    ]
    for row in samples:
        ws.append(row)
    # Reference sheet for dropdowns
    ref = wb.create_sheet('Reference')
    ref.sheet_state = 'hidden'
    sources = Source.query.order_by(Source.name).all()
    src_names = [s.name for s in sources] or ['Referral', 'WhatsApp', 'Website', 'Walk-in', 'Social Media', 'Other']
    for i, s in enumerate(src_names, 1):
        ref.cell(i, 1, s)
    types = ['Individual', 'Company', 'Investor']
    for i, t in enumerate(types, 1):
        ref.cell(i, 2, t)
    # Get staff names for Primary Representative dropdown
    users = User.query.filter_by(active=True).order_by(User.name).all()
    staff_names = [u.name for u in users]
    for i, name in enumerate(staff_names, 1):
        ref.cell(i, 3, name)
    # Source dropdown — column J (10)
    src_range = f'Reference!$A$1:$A${len(src_names)}'
    dv_src = DataValidation(type='list', formula1=src_range, allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_src)
    dv_src.add('J2:J1000')
    # Type dropdown — column I (9)
    dv_type = DataValidation(type='list', formula1='Reference!$B$1:$B$3', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_type)
    dv_type.add('I2:I1000')
    # Primary Representative dropdown — column K (11)
    staff_range = f'Reference!$C$1:$C${len(staff_names)}'
    dv_staff = DataValidation(type='list', formula1=staff_range, allow_blank=False, showDropDown=False)
    ws.add_data_validation(dv_staff)
    dv_staff.add('K2:K1000')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='customer_import_template.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/customers/import', methods=['GET', 'POST'])
@login_required
@admin_required
def import_customers():
    sources = Source.query.order_by(Source.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['sales','operations','admin'])).all()
    if request.method == 'POST':
        f = request.files.get('file')
        if not f or not f.filename.endswith('.xlsx'):
            flash('Please upload an .xlsx file')
            return redirect(url_for('import_customers'))
        from openpyxl import load_workbook
        wb = load_workbook(f)
        ws = wb.active
        imported = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            name = str(row[0]).strip()
            if not name: continue
            phone = str(row[1]).strip() if row[1] else ''
            company = str(row[2]).strip() if row[2] else ''
            phone2 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            email = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            address = str(row[5]).strip() if len(row) > 5 and row[5] else ''
            source = str(row[6]).strip() if len(row) > 6 and row[6] else ''
            nationality = str(row[7]).strip() if len(row) > 7 and row[7] else ''
            ctype = str(row[8]).strip() if len(row) > 8 and row[8] else 'Individual'
            notes = str(row[9]).strip() if len(row) > 9 and row[9] else ''
            c = Customer(name=name, phone=phone, company=company, phone2=phone2 or None,
                        email=email, address=address, source=source,
                        nationality=nationality or None, customer_type=ctype, notes=notes)
            db.session.add(c)
            imported += 1
        db.session.commit()
        flash(f'Import complete: {imported} customers added')
        return redirect(url_for('customers'))
    return render_template('import_customers.html', sources=sources, users=users)

# ── Jobs ──────────────────────────────────────────────────────────────────────

JOB_STATUSES = ['Assigned', 'Job Started', 'Processing', 'Pending Authority', 'On Hold', 'Delayed', 'Final Stage', 'Done']
JOB_STATUSES_FINANCE = ['Closed']  # Finance-only status
JOB_STATUSES_ALL = ['Pending Finance Approval'] + JOB_STATUSES + ['Pending Finance Close', 'Closed']

@app.route('/jobs')
@login_required
def jobs():
    now = now_dubai()
    role = session['role']
    sort = request.args.get('sort', 'due')
    order = request.args.get('order', 'asc')
    status_filter = request.args.get('status', '')
    
    try:
        # Exclude Done and Closed by default unless explicitly filtered
        if not status_filter:
            if role in ['admin', 'finance']:
                job_list = Job.query.all()
            else:
                job_list = Job.query.filter(Job.status.notin_(['Done', 'Closed', 'Closed - Pending Partner Commission'])).order_by(Job.due_date.asc()).all()
        elif status_filter == 'Closed':
            job_list = Job.query.filter(Job.status.in_(['Closed', 'Closed - Pending Partner Commission'])).all()
        elif status_filter == 'Done':
            job_list = Job.query.filter(Job.status == 'Done').all()
        else:
            job_list = Job.query.filter(Job.status == status_filter).order_by(Job.due_date.asc()).all()
        
        priority_filter = request.args.get('priority', '')
        assigned_filter = request.args.get('assigned_to', '') or request.args.get('staff', '')
        date_filter = request.args.get('date', '')
        from_date = request.args.get('from_date', '')
        to_date = request.args.get('to_date', '')

        customer_search = request.args.get('customer', '').strip().lower()
        if customer_search:
            job_list = [j for j in job_list if customer_search in (j.customer.name or '').lower() or customer_search in (j.customer.company or '').lower()]
        if status_filter and status_filter not in ["Closed", "Done"]:
            job_list = [j for j in job_list if j.status == status_filter]
        if priority_filter:
            job_list = [j for j in job_list if j.priority == priority_filter]
        if assigned_filter:
            try:
                job_list = [j for j in job_list if j.assigned_to == int(assigned_filter)]
            except: pass
        representative_filter = request.args.get('representative', '')
        if representative_filter:
            try:
                job_list = [j for j in job_list if j.customer and j.customer.assigned_to == int(representative_filter)]
            except: pass
        # Due date filters
        if date_filter == 'today':
            job_list = [j for j in job_list if j.due_date and j.due_date.date() == now.date()]
        elif date_filter == 'week':
            week_end = now + timedelta(days=7)
            job_list = [j for j in job_list if j.due_date and now.date() <= j.due_date.date() <= week_end.date()]
        elif date_filter == 'month':
            job_list = [j for j in job_list if j.due_date and j.due_date.month == now.month and j.due_date.year == now.year]
        elif date_filter == 'custom':
            if from_date:
                try:
                    fd = datetime.strptime(from_date, '%Y-%m-%d').date()
                    job_list = [j for j in job_list if j.due_date and j.due_date.date() >= fd]
                except: pass
            if to_date:
                try:
                    td = datetime.strptime(to_date, '%Y-%m-%d').date()
                    job_list = [j for j in job_list if j.due_date and j.due_date.date() <= td]
                except: pass
        overdue = [j for j in job_list if j.due_date and j.due_date < now and j.status not in ['Done', 'Pending Finance Approval']]
        users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
        jobs_invoiced = sum((j.amount_invoiced or 0) for j in job_list)
        jobs_received = sum((j.amount_received or 0) for j in job_list)
        jobs_pending = jobs_invoiced - jobs_received
        jobs_completed = sum((j.amount_received or 0) for j in job_list if j.status == 'Done')
    except Exception as e:
        # DB migration not complete yet — run it now
        try:
            with db.engine.connect() as conn:
                for col, typ in [
                    ('amount_invoiced', 'FLOAT DEFAULT 0'),
                    ('amount_received', 'FLOAT DEFAULT 0'),
                    ('num_persons', 'INTEGER DEFAULT 1'),
                    ('finance_approved_by', 'INTEGER'),
                    ('finance_approved_at', 'TIMESTAMP'),
                    ('finance_notes', 'TEXT'),
                ]:
                    try:
                        conn.execute(db.text(f'ALTER TABLE job ADD COLUMN IF NOT EXISTS {col} {typ}'))
                    except:
                        pass
                conn.commit()
        except:
            pass
        flash('System update applied. Please refresh.')
        return redirect(url_for('dashboard'))
    return render_template('jobs.html', jobs=job_list, now=now, overdue=overdue,
                           statuses=JOB_STATUSES + (['Closed'] if session.get('role') in ['admin','finance'] else []), users=users,
                           status_filter=status_filter, priority_filter=priority_filter,
                           assigned_filter=assigned_filter, date_filter=date_filter,
                           sort=sort, order=order,
                           jobs_invoiced=jobs_invoiced, jobs_received=jobs_received,
                           jobs_pending=jobs_pending, jobs_completed=jobs_completed)


@app.route('/jobs/export')
@login_required
def export_jobs():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tasks'
    headers = ['ID','Customer','Company','Service Type','Assigned To','Status','Priority','Due Date','Invoiced (AED)','Received (AED)','Pending (AED)','Created']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    jobs = Job.query.order_by(Job.due_date.asc()).all()
    for j in jobs:
        ws.append([
            j.id,
            j.customer.name if j.customer else '',
            j.customer.company if j.customer and j.customer.company else '',
            j.job_type or '',
            j.assignee.name if j.assignee else '',
            j.status or '',
            j.priority or '',
            j.due_date.strftime('%d/%m/%Y') if j.due_date else '',
            j.amount_invoiced or 0,
            j.amount_received or 0,
            (j.amount_invoiced or 0) - (j.amount_received or 0),
            j.created_at.strftime('%d/%m/%Y') if j.created_at else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_tasks.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/jobs/add', methods=['GET', 'POST'])
@login_required
def add_job():
    if session.get('role') not in ['admin', 'operations']:
        flash('Access denied — only Operations can add tasks')
        return redirect(url_for('jobs'))
    customers = Customer.query.order_by(Customer.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    if request.method == 'POST':
        due = request.form.get('due_date')
        due_dt = datetime.strptime(due, '%Y-%m-%d') if due else None
        amount_invoiced = request.form.get('amount_invoiced') or 0
        assigned = request.form.get('assigned_to')
        job = Job(
            customer_id=int(request.form['customer_id']),
            job_type=request.form['job_type'],
            assigned_to=int(assigned) if assigned else None,
            due_date=due_dt,
            priority=request.form.get('priority', 'Medium'),
            internal_notes=request.form.get('internal_notes'),
            service_note=request.form.get('service_note', '').strip() or None,
            amount_invoiced=float(amount_invoiced),
            amount_received=0,
            num_persons=int(request.form.get('num_persons') or 1),
            created_by=session['user_id'],
            status='Pending Finance Approval'
        )
        db.session.add(job)
        db.session.commit()
        update = JobUpdate(job_id=job.id, status='Pending Finance Approval',
                           remark='Task created — awaiting finance approval',
                           staff_name=session['user_name'])
        db.session.add(update)
        # Process sub-tasks submitted inline
        st_titles = request.form.getlist('st_title[]')
        st_service_types = request.form.getlist('st_service_type[]')
        st_assigned_tos = request.form.getlist('st_assigned_to[]')
        st_due_dates = request.form.getlist('st_due_date[]')
        st_priorities = request.form.getlist('st_priority[]')
        try:
          for i, title in enumerate(st_titles):
            if not title.strip():
                continue
            st_assigned = st_assigned_tos[i] if i < len(st_assigned_tos) and st_assigned_tos[i] else None
            st_due_str = st_due_dates[i] if i < len(st_due_dates) and st_due_dates[i] else None
            st_due = datetime.strptime(st_due_str, '%Y-%m-%d') if st_due_str else now_dubai() + timedelta(days=1)
            subtask = SubTask(
                job_id=job.id,
                title=title.strip(),
                service_type=st_service_types[i] if i < len(st_service_types) else None,
                assigned_to=int(st_assigned) if st_assigned else (job.assigned_to or session['user_id']),
                due_date=st_due,
                priority=st_priorities[i] if i < len(st_priorities) else 'Medium',
            )
            db.session.add(subtask)
          db.session.commit()
        except Exception as e:
          db.session.rollback()
          print(f'SubTask error: {e}')

        # Handle additional tasks for same customer
        extra_types = request.form.getlist('extra_job_type[]')
        extra_assigned = request.form.getlist('extra_assigned_to[]')
        extra_due = request.form.getlist('extra_due_date[]')
        extra_priority = request.form.getlist('extra_priority[]')
        extra_amount = request.form.getlist('extra_amount[]')
        extra_persons = request.form.getlist('extra_persons[]')
        extra_notes = request.form.getlist('extra_notes[]')
        extra_service_notes = request.form.getlist('extra_service_note[]')

        for i, jt in enumerate(extra_types):
            if not jt: continue
            try: ea = int(extra_assigned[i]) if i < len(extra_assigned) and extra_assigned[i] else None
            except: ea = None
            try: ed = datetime.strptime(extra_due[i], '%Y-%m-%d') if i < len(extra_due) and extra_due[i] else now_dubai() + timedelta(days=1)
            except: ed = now_dubai() + timedelta(days=1)
            try: eamt = float(extra_amount[i]) if i < len(extra_amount) and extra_amount[i] else 0
            except: eamt = 0
            try: ep = int(extra_persons[i]) if i < len(extra_persons) and extra_persons[i] else 1
            except: ep = 1
            extra_job = Job(
                customer_id=job.customer_id,
                job_type=jt,
                assigned_to=ea,
                due_date=ed,
                priority=extra_priority[i] if i < len(extra_priority) else 'Medium',
                amount_invoiced=eamt,
                num_persons=ep,
                internal_notes=extra_notes[i] if i < len(extra_notes) else None,
                service_note=extra_service_notes[i].strip() if i < len(extra_service_notes) and extra_service_notes[i].strip() else None,
                status='Pending Finance Approval',
                created_by=session['user_id']
            )
            db.session.add(extra_job)

        db.session.commit()
        count = 1 + len([t for t in extra_types if t])
        flash(f'{count} task(s) created successfully')
        return redirect(url_for('jobs'))
    tomorrow = (now_dubai() + timedelta(days=1)).strftime('%Y-%m-%d')
    import json
    service_days = {jt.name: (getattr(jt, 'default_days', None) or 1) for jt in job_types}
    all_jobs = Job.query.order_by(Job.created_at.desc()).all()
    return render_template('add_job.html', customers=customers, job_types=job_types, users=users, tomorrow=tomorrow, service_days=json.dumps(service_days), all_jobs=all_jobs)

@app.route('/jobs/<int:job_id>', methods=['GET', 'POST'])
@login_required
def job_detail(job_id):
    job = Job.query.get_or_404(job_id)
    now = now_dubai()
    role = session['role']
    # Sales and operations can view all tasks (not just assigned ones)
    # Only restrict if somehow a non-authorised role gets here
    if role not in ['admin', 'sales', 'operations', 'finance', 'staff']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    if request.method == 'POST':
        # Sales cannot update tasks at all
        if role == 'sales':
            flash('Sales cannot update task status. Contact Operations.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Closed — no updates except admin/finance
        if job.status == 'Closed' and role not in ['admin', 'finance']:
            flash('This task is closed. No further updates allowed.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Done/Pending Finance Close — no further updates from non-admin/finance
        if job.status in ['Done', 'Pending Finance Close'] and role not in ['admin', 'finance']:
            flash('Task is already marked Done. Contact Finance/Admin for changes.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Block sales/staff from updating if pending finance approval
        if job.status == 'Pending Finance Approval' and role in ['staff', 'sales']:
            flash('This task is pending finance approval. You cannot update it yet.')
            return redirect(url_for('job_detail', job_id=job_id))
        # Block sales/staff from updating if pending finance close
        if job.status == 'Pending Finance Close' and role in ['staff', 'sales']:
            flash('Work is complete. Awaiting finance to close this task.')
            return redirect(url_for('job_detail', job_id=job_id))
        remark = request.form.get('remark', '').strip()
        if not remark:
            flash('Remark is required')
            return redirect(url_for('job_detail', job_id=job_id))
        new_status = request.form.get('status', job.status)
        if role == 'staff' and new_status == 'Pending Finance Approval':
            new_status = job.status
        # When ops marks Done → stays as Done, appears in Finance queue
        # Finance will verify payment and close the task
        if new_status == 'Done' and role not in ['admin', 'finance']:
            pass  # Keep as Done — Finance will close it
        job.status = new_status
        # Save completion fields when marking Done or Pending Finance Close
        if new_status in ['Done', 'Pending Finance Close']:
            if not job.completed_at:
                job.completed_at = now_dubai()
            job.final_remarks = request.form.get('final_remarks') or None
            job.future_work_notes = request.form.get('future_work_notes') or None
            # Log completion to timeline
            completion_note = 'Task completed.'
            if job.final_remarks: completion_note += f' Remarks: {job.final_remarks}'
            update_completion = JobUpdate(job_id=job.id, status=new_status, remark=completion_note, staff_name=session['user_name'])
            db.session.add(update_completion)
        update = JobUpdate(job_id=job.id, status=new_status,
                           status_note=request.form.get('status_note', '').strip()[:100] or None,
                           remark=remark, staff_name=session['user_name'])
        db.session.add(update)
        db.session.commit()
        flash('Task updated')
        return redirect(url_for('job_detail', job_id=job_id))
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    service_types = ServiceType.query.order_by(ServiceType.name).all()
    # All jobs for same customer (for multi-task timeline)
    sibling_jobs = Job.query.filter_by(customer_id=job.customer_id).order_by(Job.created_at.asc()).all()
    partners = Partner.query.filter_by(active=True).order_by(Partner.name).all()
    return render_template('job_detail.html', job=job, now=now,
                           statuses=JOB_STATUSES, users=users,
                           service_types=service_types, timedelta=timedelta,
                           sibling_jobs=sibling_jobs, partners=partners)

@app.route('/jobs/<int:job_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_job(job_id):
    job = Job.query.get_or_404(job_id)
    if session.get('role') not in ['admin', 'operations']:
        flash('Access denied — only Operations can edit tasks')
        return redirect(url_for('job_detail', job_id=job_id))
    if job.status == 'Closed' and session['role'] != 'admin':
        flash('Closed tasks cannot be edited.')
        return redirect(url_for('job_detail', job_id=job_id))
    # Staff can only edit tasks assigned to them
    if session['role'] in ['sales', 'staff'] and job.assigned_to != session['user_id']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    customers = Customer.query.order_by(Customer.name).all()
    job_types = ServiceType.query.order_by(ServiceType.name).all()
    users = User.query.filter_by(active=True).filter(User.role.in_(['staff', 'sales', 'operations', 'admin'])).all()
    if request.method == 'POST':
        job.job_type = request.form['job_type']
        job.customer_id = int(request.form['customer_id'])
        assigned = request.form.get('assigned_to')
        job.assigned_to = int(assigned) if assigned else None
        due = request.form.get('due_date')
        job.due_date = datetime.strptime(due, '%Y-%m-%d') if due else None
        job.priority = request.form.get('priority', 'Medium')
        job.internal_notes = request.form.get('internal_notes')
        job.service_note = request.form.get('service_note', '').strip() or None
        if request.form.get('num_persons'):
            job.num_persons = int(request.form.get('num_persons'))
        try:
            ai = request.form.get('amount_invoiced')
            ar = request.form.get('amount_received')
            if ai: job.amount_invoiced = float(ai)
            if ar: job.amount_received = float(ar)
        except:
            pass
        db.session.commit()
        flash('Task updated')
        return redirect(url_for('job_detail', job_id=job_id))
    return render_template('edit_job.html', job=job, customers=customers,
                           job_types=job_types, users=users, statuses=JOB_STATUSES)

@app.route('/jobs/<int:job_id>/delete')
@login_required
def delete_job(job_id):
    if session['role'] not in ['admin', 'operations']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    job = Job.query.get_or_404(job_id)
    SubTask.query.filter_by(job_id=job_id).delete()
    JobUpdate.query.filter_by(job_id=job_id).delete()
    db.session.delete(job)
    db.session.commit()
    flash('Task deleted')
    return redirect(url_for('jobs'))

@app.route('/jobs/bulk-delete', methods=['POST'])
@login_required
def bulk_delete_jobs():
    if session['role'] not in ['admin', 'operations']:
        flash('Access denied')
        return redirect(url_for('jobs'))
    ids = request.form.getlist('job_ids')
    if not ids:
        flash('No tasks selected')
        return redirect(url_for('jobs'))
    count = 0
    for job_id in ids:
        job = Job.query.get(int(job_id))
        if job:
            SubTask.query.filter_by(job_id=job.id).delete()
            JobUpdate.query.filter_by(job_id=job.id).delete()
            db.session.delete(job)
            count += 1
    db.session.commit()
    flash(f'{count} task(s) deleted')
    return redirect(url_for('jobs'))

# ── Finance ───────────────────────────────────────────────────────────────────

@app.route('/jobs/<int:job_id>/approve', methods=['POST'])
@login_required
@finance_required
def approve_job(job_id):
    job = Job.query.get_or_404(job_id)
    amount_invoiced = request.form.get('amount_invoiced', '').strip()
    amount_received = request.form.get('amount_received', '').strip()
    try:
        if amount_invoiced:
            job.amount_invoiced = float(amount_invoiced)
        if amount_received:
            job.amount_received = float(amount_received)
    except:
        pass
    # If task was Done, Finance is closing it; if Pending Finance Approval, Finance is approving it
    if job.status == 'Done':
        job.status = 'Closed'
        job.completed_at = now_dubai()
    else:
        job.status = 'Assigned'
    job.finance_approved_by = session['user_id']
    job.finance_approved_at = now_dubai()
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        job.finance_notes = notes  # save to job record
    action = 'Closed' if job.status == 'Closed' else 'Approved'
    remark = f'{action} by Finance. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f}'
    if notes:
        remark += f'. Notes: {notes}'
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    msg = 'Task closed successfully.' if job.status == 'Closed' else 'Task approved and assigned to staff.'
    flash(msg)
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/jobs/<int:job_id>/payment', methods=['POST'])
@login_required
@finance_required
def update_payment(job_id):
    job = Job.query.get_or_404(job_id)
    try:
        job.amount_invoiced = float(request.form.get('amount_invoiced') or job.amount_invoiced or 0)
        job.amount_received = float(request.form.get('amount_received') or job.amount_received or 0)
    except:
        pass
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        # Replace finance notes (don't append)
        job.finance_notes = notes
    remark = f'Payment updated. Invoiced: AED {job.amount_invoiced:,.0f} / Received: AED {job.amount_received:,.0f}'
    if notes:
        remark += f'. Notes: {notes}'
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    flash('Payment updated.')
    return redirect(request.referrer or url_for('jobs'))


@app.route('/jobs/<int:job_id>/close', methods=['POST'])
@login_required
@finance_required
def close_job(job_id):
    job = Job.query.get_or_404(job_id)
    
    # Update invoice and received amounts
    try:
        ai = request.form.get('amount_invoiced')
        ar = request.form.get('amount_received')
        if ai: job.amount_invoiced = float(ai)
        if ar: job.amount_received = float(ar)
    except:
        pass
    
    # Handle partner commission choice (mandatory)
    partner_choice = request.form.get('partner_commission_expected')
    
    if partner_choice == 'no':
        # REGULAR TASK - Revenue counted immediately
        try:
            rev = request.form.get('revenue')
            if rev:
                job.revenue = float(rev)
                job.revenue_date = now_dubai().date()  # Revenue counted today (cash-basis)
            else:
                flash('Revenue is required for regular tasks.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
        except:
            flash('Invalid revenue amount.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        job.partner_commission_expected = False
        job.partner_name = None
        job.partner_amount = None
        job.partner_due_date = None
        job.partner_status = None
        job.status = 'Closed'
        
        remark = f'Task CLOSED by Finance. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Revenue: AED {job.revenue:,.0f} (counted for {now_dubai().strftime("%B %Y")})'
        
    elif partner_choice == 'yes':
        # PARTNER COMMISSION TASK - Revenue = 0 until partner pays
        partner_name = request.form.get('partner_name')
        new_partner_name = request.form.get('new_partner_name', '').strip()
        
        # Handle new partner creation
        if partner_name == '__ADD_NEW__':
            if not new_partner_name:
                flash('Please enter new partner name.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
            # Create new partner
            try:
                new_partner = Partner(name=new_partner_name)
                db.session.add(new_partner)
                db.session.flush()  # Get the ID without committing
                partner_name = new_partner_name
            except:
                flash('Partner name already exists or invalid.', 'error')
                return redirect(url_for('job_detail', job_id=job_id))
        
        if not partner_name or partner_name == '__ADD_NEW__':
            flash('Please select a partner.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        try:
            partner_amount = float(request.form.get('partner_amount'))
            partner_due_date = request.form.get('partner_due_date')
            if not partner_due_date:
                raise ValueError("Due date required")
            partner_due_date = datetime.strptime(partner_due_date, '%Y-%m-%d').date()
        except:
            flash('Partner commission amount and due date are required.', 'error')
            return redirect(url_for('job_detail', job_id=job_id))
        
        job.partner_commission_expected = True
        job.partner_name = partner_name
        job.partner_amount = partner_amount
        job.partner_due_date = partner_due_date
        job.partner_status = 'Pending'
        job.revenue = 0  # Revenue NOT counted yet
        job.status = 'Closed - Pending Partner Commission'
        
        remark = f'Task CLOSED by Finance. Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Partner Commission: {partner_name} - AED {partner_amount:,.0f} (pending). Revenue will be counted when partner pays.'
    
    else:
        flash('Please select whether partner commission is expected.', 'error')
        return redirect(url_for('job_detail', job_id=job_id))
    
    # Finance notes
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        job.finance_notes = notes
        remark += f' Notes: {notes}'
    
    # Create update record
    update = JobUpdate(job_id=job.id, status=job.status, remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    
    flash('Task closed successfully.')
    return redirect(url_for('dashboard'))

@app.route('/jobs/<int:job_id>/edit_finance', methods=['POST'])
@login_required
@finance_required
def edit_finance(job_id):
    job = Job.query.get_or_404(job_id)
    if job.status != 'Closed':
        flash('Can only edit finance details for closed tasks.')
        return redirect(url_for('job_detail', job_id=job_id))
    
    old_invoiced = job.amount_invoiced or 0
    old_received = job.amount_received or 0
    old_revenue = job.revenue or 0
    
    try:
        ai = request.form.get('amount_invoiced')
        ar = request.form.get('amount_received')
        rev = request.form.get('revenue')
        if ai: job.amount_invoiced = float(ai)
        if ar: job.amount_received = float(ar)
        if rev:
            job.revenue = float(rev)
            # Update revenue_date when revenue is edited
            if not job.revenue_date:
                job.revenue_date = now_dubai().date()
    except:
        flash('Invalid finance values.')
        return redirect(url_for('job_detail', job_id=job_id))
    
    notes = request.form.get('finance_notes', '').strip()
    if notes:
        # Replace finance notes (don't append)
        job.finance_notes = notes
    
    remark = f'Finance details EDITED by {session["user_name"]}. Previous — Invoiced: AED {old_invoiced:,.0f} / Received: AED {old_received:,.0f} / Revenue: AED {old_revenue:,.0f}. Updated — Invoiced: AED {job.amount_invoiced or 0:,.0f} / Received: AED {job.amount_received or 0:,.0f} / Revenue: AED {job.revenue or 0:,.0f}'
    if old_revenue_date != job.revenue_date:
        remark += f' / Revenue Date: {job.revenue_date.strftime("%d-%b-%Y") if job.revenue_date else "None"}'
    if old_created_at.date() != job.created_at.date():
        remark += f' / Created: {job.created_at.strftime("%d-%b-%Y")}'
    
    if notes:
        remark += f'. Notes: {notes}'
    
    update = JobUpdate(job_id=job.id, status='Closed', remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    flash('Finance details updated successfully.')
    return redirect(url_for('job_detail', job_id=job_id))

@app.route('/jobs/<int:job_id>/partial_revenue/add', methods=['POST'])
@login_required
def add_partial_revenue(job_id):
    if session['role'] not in ['finance', 'admin']:
        flash('Only Finance can record partial revenue')
        return redirect(url_for('job_detail', job_id=job_id))
    
    job = Job.query.get_or_404(job_id)
    amount = request.form.get('amount', type=float)
    revenue_date_str = request.form.get('revenue_date')
    notes = request.form.get('notes', '').strip()
    
    if not amount or amount <= 0:
        flash('Please enter a valid amount')
        return redirect(url_for('job_detail', job_id=job_id))
    
    if not revenue_date_str:
        flash('Please select a revenue date')
        return redirect(url_for('job_detail', job_id=job_id))
    
    try:
        revenue_date = datetime.strptime(revenue_date_str, '%Y-%m-%d').date()
    except:
        flash('Invalid date format')
        return redirect(url_for('job_detail', job_id=job_id))
    
    # Check total partial revenue doesn't exceed received amount
    existing_partial = sum(pr.amount for pr in job.partial_revenues)
    if existing_partial + amount > (job.amount_received or 0):
        flash(f'Total partial revenue ({existing_partial + amount:,.0f}) cannot exceed received amount ({job.amount_received:,.0f})')
        return redirect(url_for('job_detail', job_id=job_id))
    
    partial = PartialRevenue(
        job_id=job_id,
        amount=amount,
        revenue_date=revenue_date,
        notes=notes,
        recorded_by=session['user_id']
    )
    db.session.add(partial)
    db.session.commit()
    
    flash(f'Partial revenue of AED {amount:,.0f} recorded successfully')
    return redirect(url_for('job_detail', job_id=job_id))

# ── Daily Activity Log ────────────────────────────────────────────────────────

# ACTIVITIES loaded from DB — see get_activities()
ACTIVITY_DEFAULTS = [
    ('calls_existing',       'Calls to Existing/Potential Clients', 30),
    ('calls_cold',           'Cold Calling to Customer List',       30),
    ('dm_instagram',         'Instagram Direct Messages',           30),
    ('dm_facebook',          'Facebook Messages',                   30),
    ('dm_linkedin',          'LinkedIn Messages',                   30),
    ('posts_social',         'Social Media Posts (IG/FB/LinkedIn)', 2),
    ('videos_instagram',     'Instagram Video (Cross-post)',        1),
    ('linkedin_writing',     'LinkedIn Writing/Articles',           1),
    ('whatsapp_prospecting', 'WhatsApp Prospecting',               30),
    ('community_active',     'Active in Communities',               2),
    ('google_reviews',       'Google Review Collection',            6),
    ('real_estate_relations','Real Estate Agent Relationships',     2),
    ('content_marketing',    'Content for Marketing',               2),
    ('referral_building',    'Referral Building',                   2),
    ('networking_activities','Networking/Community Activities',     1),
    ('networking_events',    'Attended Networking Event',           1),
]

def get_activities():
    try:
        types = ActivityType.query.filter_by(active=True).order_by(ActivityType.sort_order, ActivityType.id).all()
        return [(t.field_key, t.label, t.weekly_target) for t in types]
    except:
        return ACTIVITY_DEFAULTS

@app.route('/activity')
@login_required
def activity_log():
    if session['role'] not in ['sales', 'admin']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    now = now_dubai()
    week_start = (now - timedelta(days=now.weekday())).date()
    from_date = request.args.get('from', week_start.strftime('%Y-%m-%d'))
    to_date = request.args.get('to', now.date().strftime('%Y-%m-%d'))
    view = request.args.get('view', 'week')

    from_dt = datetime.strptime(from_date, '%Y-%m-%d').date()
    to_dt = datetime.strptime(to_date, '%Y-%m-%d').date()

    try:
        if session['role'] == 'admin':
            sales_users = User.query.filter(User.role == 'sales', User.active == True).all()
            logs = ActivityLog.query.filter(
                ActivityLog.log_date >= from_dt,
                ActivityLog.log_date <= to_dt
            ).all()
        else:
            sales_users = [User.query.get(session['user_id'])]
            logs = ActivityLog.query.filter_by(user_id=session['user_id']).filter(
                ActivityLog.log_date >= from_dt,
                ActivityLog.log_date <= to_dt
            ).all()
    except Exception as e:
        # Table may not exist yet — use empty data
        logs = []
        sales_users = [] if session['role'] == 'admin' else [User.query.get(session['user_id'])]
        try:
            if session['role'] == 'admin':
                sales_users = User.query.filter(User.role == 'sales', User.active == True).all()
        except:
            pass

    # Build summary per user
    user_summaries = {}
    for u in sales_users:
        user_logs = [l for l in logs if l.user_id == u.id]
        summary = {}
        for field, label, target in get_activities():
            total = sum(getattr(l, field, 0) or 0 for l in user_logs)
            days = (to_dt - from_dt).days + 1
            weeks = max(1, days / 6)  # 6-day UAE working week (Sat–Thu)
            period_target = round(target * weeks)
            pct = round((total / period_target * 100) if period_target > 0 else 0)
            summary[field] = {'total': total, 'target': period_target, 'pct': pct}
        user_summaries[u.id] = {'user': u, 'summary': summary, 'logs': user_logs}

    # Today's log for current user (for the entry form)
    today_log = None
    if session['role'] == 'sales':
        today_log = ActivityLog.query.filter_by(
            user_id=session['user_id'],
            log_date=now.date()
        ).first()

    try:
        activity_types = ActivityType.query.filter_by(active=True).order_by(ActivityType.sort_order, ActivityType.id).all()
    except Exception:
        activity_types = []
    return render_template('activity_log.html',
                           activities=get_activities(),
                           activity_types=activity_types,
                           user_summaries=user_summaries,
                           sales_users=sales_users,
                           today_log=today_log,
                           from_date=from_date, to_date=to_date,
                           view=view, now=now)

@app.route('/activity/log', methods=['POST'])
@login_required
def save_activity():
    if session['role'] not in ['sales', 'admin']:
        flash('Access denied')
        return redirect(url_for('dashboard'))
    log_date_str = request.form.get('log_date', now_dubai().date().strftime('%Y-%m-%d'))
    log_date = datetime.strptime(log_date_str, '%Y-%m-%d').date()
    # Admin can log for any user
    user_id = int(request.form.get('user_id_override') or session['user_id'])
    if request.form.get('user_id_override') and session['role'] != 'admin':
        flash('Access denied'); return redirect(url_for('activity_log'))

    # Upsert — update if exists for this date
    log = ActivityLog.query.filter_by(user_id=user_id, log_date=log_date).first()
    if not log:
        log = ActivityLog(user_id=user_id, log_date=log_date)
        db.session.add(log)

    for field, label, target in get_activities():
        val = request.form.get(field, '0').strip()
        try:
            setattr(log, field, int(val) if val else 0)
        except:
            setattr(log, field, 0)
    log.off_day = request.form.get('off_day', '') or None
    log.notes = request.form.get('notes', '')
    log.updated_at = now_dubai()
    db.session.commit()
    flash(f'Activity log saved for {log_date.strftime("%d %b %Y")}')
    return redirect(url_for('activity_log'))



@app.route('/activity/<int:log_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_activity_log(log_id):
    log = ActivityLog.query.get_or_404(log_id)
    if session['role'] != 'admin' and log.user_id != session['user_id']:
        flash('Access denied')
        return redirect(url_for('activity_log'))
    if request.method == 'POST':
        for field, label, target in get_activities():
            val = request.form.get(field, '0').strip()
            try: setattr(log, field, int(val) if val else 0)
            except: setattr(log, field, 0)
        log.off_day = request.form.get('off_day', '') or None
        log.notes = request.form.get('notes', '')
        log.updated_at = now_dubai()
        db.session.commit()
        flash(f'Activity log updated for {log.log_date.strftime("%d %b %Y")}')
        return redirect(url_for('activity_log'))
    return redirect(url_for('activity_log'))

@app.route('/activity/<int:log_id>/delete')
@login_required
def delete_activity_log(log_id):
    log = ActivityLog.query.get_or_404(log_id)
    # Admin can delete any, sales can delete their own
    if session['role'] != 'admin' and log.user_id != session['user_id']:
        flash('Access denied')
        return redirect(url_for('activity_log'))
    db.session.delete(log)
    db.session.commit()
    flash('Activity log entry deleted')
    return redirect(url_for('activity_log'))


# ── Admin — Activity Types ────────────────────────────────────────────────────

@app.route('/admin/activity-type/add', methods=['POST'])
@login_required
@admin_required
def admin_add_activity_type():
    label = request.form.get('label', '').strip()
    target = request.form.get('daily_target', '1').strip()
    if not label:
        flash('Activity name is required')
        return redirect(url_for('activity_log'))
    # Generate a safe field_key from label
    import re as re_mod
    field_key = re_mod.sub(r'[^a-z0-9]', '_', label.lower())[:40]
    field_key = re_mod.sub(r'_+', '_', field_key).strip('_')
    # Ensure unique
    base_key = field_key
    counter = 1
    while ActivityType.query.filter_by(field_key=field_key).first():
        field_key = f'{base_key}_{counter}'
        counter += 1
    try:
        target_val = float(target)
    except:
        target_val = 1.0
    max_order = db.session.query(db.func.max(ActivityType.sort_order)).scalar() or 0
    at = ActivityType(field_key=field_key, label=label, weekly_target=target_val, sort_order=max_order+1)
    db.session.add(at)
    db.session.commit()
    flash(f'Activity "{label}" added')
    return redirect(url_for('activity_log'))

@app.route('/admin/activity-type/<int:type_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_activity_type(type_id):
    at = ActivityType.query.get_or_404(type_id)
    at.label = request.form.get('label', at.label).strip()
    try:
        at.weekly_target = float(request.form.get('weekly_target', at.weekly_target))
    except:
        pass
    db.session.commit()
    flash(f'Activity updated')
    return redirect(url_for('activity_log'))

@app.route('/admin/activity-type/<int:type_id>/delete')
@login_required
@admin_required
def admin_delete_activity_type(type_id):
    at = ActivityType.query.get_or_404(type_id)
    at.active = False  # Soft delete — preserve historical data
    db.session.commit()
    flash(f'Activity "{at.label}" removed')
    return redirect(url_for('activity_log'))


# ── Admin Edit Routes ─────────────────────────────────────────────────────────

@app.route('/admin/service/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_service(item_id):
    item = Service.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Service updated')
    return redirect(url_for('admin_panel') + '#services')

@app.route('/admin/source/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_source(item_id):
    item = Source.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Source updated')
    return redirect(url_for('admin_panel') + '#sources')

@app.route('/admin/campaign/add', methods=['POST'])
@login_required
@admin_required
def admin_add_campaign():
    name = request.form.get('name', '').strip()
    if name:
        existing = Campaign.query.filter_by(name=name).first()
        if existing:
            flash('Campaign already exists')
        else:
            db.session.add(Campaign(name=name))
            db.session.commit()
            flash(f'Campaign "{name}" added')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/campaign/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_campaign(item_id):
    item = Campaign.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Campaign updated')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/campaign/<int:item_id>/delete')
@login_required
@admin_required
def admin_delete_campaign(item_id):
    item = Campaign.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash(f'Campaign "{item.name}" removed')
    return redirect(url_for('admin_panel') + '#campaigns')

@app.route('/admin/jobtype/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_jobtype(item_id):
    item = ServiceType.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
    try:
        item.default_days = int(request.form.get('default_days', 1))
    except:
        pass
    db.session.commit()
    flash('Service type updated')
    return redirect(url_for('admin_panel') + '#service-types')

@app.route('/admin/doctype/<int:item_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_doctype(item_id):
    item = DocType.query.get_or_404(item_id)
    name = request.form.get('name', '').strip()
    if name:
        item.name = name
        db.session.commit()
        flash('Document type updated')
    return redirect(url_for('admin_panel'))

# ── Admin — Job Types ─────────────────────────────────────────────────────────

@app.route('/admin/jobtype/add', methods=['POST'])
@login_required
@admin_required
def admin_add_jobtype():
    name = request.form.get('name', '').strip()
    if name:
        if not ServiceType.query.filter_by(name=name).first():
            try:
                days = int(request.form.get('default_days', 1))
            except:
                days = 1
            new_jt = ServiceType(name=name)
            new_jt.default_days = days
            db.session.add(new_jt)
            db.session.commit()
            flash(f'Service type "{name}" added')
        else:
            flash('Job type already exists')
    return redirect(url_for('admin_panel') + '#service-types')

@app.route('/admin/jobtype/<int:jobtype_id>/delete')
@login_required
@admin_required
def admin_delete_jobtype(jobtype_id):
    jt = ServiceType.query.get_or_404(jobtype_id)
    db.session.delete(jt)
    db.session.commit()
    flash(f'Job type "{jt.name}" removed')
    return redirect(url_for('admin_panel'))

# ── Documents ─────────────────────────────────────────────────────────────────

@app.route('/documents')
@login_required
def documents():
    now = now_dubai()
    search = request.args.get('search', '').strip().lower()
    belongs_filter = request.args.get('belongs_to', '')
    doc_type_filter = request.args.get('doc_type', '')
    customer_filter = request.args.get('customer_id', '')
    expiry_filter = request.args.get('expiry', '')

    try:
        doc_list = Document.query.order_by(Document.expiry_date).all()
    except Exception:
        # Run missing column migrations inline
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE document ADD COLUMN IF NOT EXISTS file_name VARCHAR(255)'))
                conn.commit()
        except Exception:
            pass
        flash('System updated. Please refresh.')
        return redirect(url_for('dashboard'))

    # Summary counts (all docs)
    total_docs = len(doc_list)
    expiring_30 = [d for d in doc_list if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30]
    expiring_60 = [d for d in doc_list if d.expiry_date and 30 < (d.expiry_date - now).days <= 60]
    expiring_90 = [d for d in doc_list if d.expiry_date and 60 < (d.expiry_date - now).days <= 90]
    expired_docs = [d for d in doc_list if d.expiry_date and d.expiry_date < now]

    # Apply filters
    if search:
        doc_list = [d for d in doc_list if
                    search in (d.owner_name or '').lower() or
                    search in (d.doc_type or '').lower() or
                    (d.customer and search in d.customer.name.lower()) or
                    (d.customer and d.customer.company and search in d.customer.company.lower())]
    if belongs_filter:
        doc_list = [d for d in doc_list if d.belongs_to == belongs_filter]
    if doc_type_filter:
        doc_list = [d for d in doc_list if d.doc_type == doc_type_filter]
    if customer_filter:
        doc_list = [d for d in doc_list if d.customer_id == int(customer_filter)]
    if expiry_filter == '30':
        doc_list = [d for d in doc_list if d.expiry_date and 0 <= (d.expiry_date - now).days <= 30]
    elif expiry_filter == '60':
        doc_list = [d for d in doc_list if d.expiry_date and 30 < (d.expiry_date - now).days <= 60]
    elif expiry_filter == '90':
        doc_list = [d for d in doc_list if d.expiry_date and 60 < (d.expiry_date - now).days <= 90]
    elif expiry_filter == 'expired':
        doc_list = [d for d in doc_list if d.expiry_date and d.expiry_date < now]

    # Pagination
    page = int(request.args.get('page', 1))
    per_page = 50
    total = len(doc_list)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    paginated = doc_list[(page-1)*per_page: page*per_page]

    customers = Customer.query.order_by(Customer.name).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    
    # Count documents per customer (for showing multiple file indicator)
    from collections import defaultdict
    customer_doc_count = defaultdict(int)
    all_docs = Document.query.all()
    for d in all_docs:
        if d.customer_id:
            customer_doc_count[d.customer_id] += 1
    
    # Get all documents grouped by customer for popup (convert to dicts for JSON)
    customer_docs = defaultdict(list)
    for d in all_docs:
        if d.customer_id and d.file_url:
            customer_docs[d.customer_id].append({
                'doc_type': d.doc_type or 'Document',
                'file_name': d.file_name or 'Unnamed file',
                'file_url': d.file_url
            })
    
    return render_template('documents.html',
                           documents=paginated, customers=customers, doc_types=doc_types,
                           total_docs=total_docs, expiring_30=len(expiring_30),
                           expiring_60=len(expiring_60), expiring_90=len(expiring_90),
                           expired_count=len(expired_docs),
                           search=search, belongs_filter=belongs_filter,
                           doc_type_filter=doc_type_filter, customer_filter=customer_filter,
                           expiry_filter=expiry_filter,
                           total=total, page=page, total_pages=total_pages,
                           now=now,
                           customer_doc_count=dict(customer_doc_count),
                           customer_docs=dict(customer_docs))


@app.route('/documents/export')
@login_required
def export_documents():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from flask import send_file
    wb = Workbook()
    ws = wb.active
    ws.title = 'Documents'
    headers = ['Customer','Company','Doc Type','Belongs To','Owner Name','Expiry Date','Days Until Expiry','Notes','Added By','Created']
    for i, h in enumerate(headers, 1):
        ws.cell(1, i, h).font = Font(bold=True, color='FFFFFF')
        ws.cell(1, i).fill = PatternFill('solid', fgColor='1A3B8B')
    docs = Document.query.order_by(Document.expiry_date).all()
    now = now_dubai()
    for d in docs:
        days = (d.expiry_date - now).days if d.expiry_date else ''
        ws.append([
            d.customer.name if d.customer else '',
            d.customer.company if d.customer and d.customer.company else '',
            d.doc_type or '',
            d.belongs_to or '',
            d.owner_name or '',
            d.expiry_date.strftime('%d/%m/%Y') if d.expiry_date else '',
            days,
            d.notes or '',
            d.added_by or '',
            d.created_at.strftime('%d/%m/%Y') if d.created_at else '',
        ])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 14)
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='tahfeel_documents.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/documents/add', methods=['GET', 'POST'])
@login_required
def add_document():
    customers = Customer.query.order_by(Customer.name).all()
    doc_types = DocType.query.order_by(DocType.name).all()
    sources = Source.query.order_by(Source.name).all()
    if request.method == 'POST':
        expiry = request.form.get('expiry_date')
        expiry_dt = datetime.strptime(expiry, '%Y-%m-%d') if expiry else None
        customer_id = request.form.get('customer_id') or None
        if request.form.get('new_customer_name'):
            new_cust = Customer(
                name=request.form['new_customer_name'],
                company=request.form.get('new_customer_company'),
                phone=request.form.get('new_customer_phone'),
                email=request.form.get('new_customer_email'),
                source=request.form.get('new_customer_source'),
            )
            db.session.add(new_cust)
            db.session.flush()
            customer_id = new_cust.id
        # Handle file upload (dummy — store filename only for now)
        file_name = None
        file_url = None
        public_id = None
        if 'document_file' in request.files:
            f = request.files['document_file']
            if f and f.filename:
                file_name = f.filename
                file_url, public_id = upload_to_cloudinary(f)
                if not file_url:
                    flash('⚠️ File could not be uploaded — document saved without attachment. Please check Cloudinary settings.', 'warning')
        doc = Document(
            doc_type=request.form['doc_type'],
            belongs_to=request.form['belongs_to'],
            owner_name=request.form['owner_name'],
            customer_id=int(customer_id) if customer_id else None,
            expiry_date=expiry_dt,
            notes=request.form.get('notes'),
            file_name=file_name,
            file_url=file_url,
            cloudinary_public_id=public_id,
            uploaded_by=session['user_id'],
            added_by=session['user_name'],
        )
        db.session.add(doc)
        db.session.commit()
        # Option A: redirect back to add form with customer pre-selected + success message
        customer_id_param = f'?customer_id={customer_id}&added=1' if customer_id else '?added=1'
        flash('Document saved successfully!')
        if request.form.get('add_another'):
            return redirect(url_for('add_document') + customer_id_param)
        elif customer_id:
            return redirect(url_for('customer_detail', customer_id=int(customer_id)))
        return redirect(url_for('documents'))
    return render_template('add_document.html', customers=customers,
                           doc_types=doc_types, sources=sources)


@app.route('/documents/<int:doc_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    doc_types = DocType.query.order_by(DocType.name).all()
    customers = Customer.query.order_by(Customer.name).all()
    if request.method == 'POST':
        doc.doc_type = request.form.get('doc_type', doc.doc_type)
        doc.belongs_to = request.form.get('belongs_to', doc.belongs_to)
        doc.owner_name = request.form.get('owner_name', doc.owner_name)
        try:
            cid = request.form.get('customer_id')
            doc.customer_id = int(cid) if cid else None
        except: pass
        try:
            ed = request.form.get('expiry_date')
            doc.expiry_date = datetime.strptime(ed, '%Y-%m-%d') if ed else None
        except: pass
        doc.notes = request.form.get('notes', doc.notes)
        # Handle new file upload
        if 'document_file' in request.files:
            f = request.files['document_file']
            if f and f.filename:
                doc.file_name = f.filename
                url, public_id = upload_to_cloudinary(f)
                if url:
                    doc.file_url = url
                    doc.cloudinary_public_id = public_id
        db.session.commit()
        flash('Document updated')
        if request.form.get('add_another'):
            cid = doc.customer_id
            return redirect(url_for('add_document') + (f'?customer_id={cid}' if cid else ''))
        return redirect(url_for('documents'))
    return render_template('edit_document.html', doc=doc, doc_types=doc_types, customers=customers)

@app.route('/documents/<int:doc_id>/delete')
@login_required
@admin_required
def delete_document(doc_id):
    doc = Document.query.get_or_404(doc_id)
    db.session.delete(doc)
    db.session.commit()
    flash('Document removed')
    return redirect(url_for('documents'))

# ── Admin — Document Types ────────────────────────────────────────────────────

@app.route('/admin/doctype/add', methods=['POST'])
@login_required
@admin_required
def admin_add_doctype():
    name = request.form.get('name', '').strip()
    if name:
        if not DocType.query.filter_by(name=name).first():
            db.session.add(DocType(name=name))
            db.session.commit()
            flash(f'Document type "{name}" added')
        else:
            flash('Document type already exists')
    return redirect(url_for('admin_panel'))

@app.route('/admin/doctype/<int:doctype_id>/delete')
@login_required
@admin_required
def admin_delete_doctype(doctype_id):
    dt = DocType.query.get_or_404(doctype_id)
    db.session.delete(dt)
    db.session.commit()
    flash(f'Document type "{dt.name}" removed')
    return redirect(url_for('admin_panel'))

# ─────────────────────────────────────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()
        # Create monthly_target table if not exists
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS monthly_target (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER REFERENCES "user"(id),
                        month INTEGER NOT NULL,
                        year INTEGER NOT NULL,
                        lead_target INTEGER DEFAULT 0,
                        conversion_target INTEGER DEFAULT 0,
                        amount_target FLOAT DEFAULT 0
                    )
                """))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS amount_target FLOAT DEFAULT 0'))
                conn.execute(db.text('ALTER TABLE customer ADD COLUMN IF NOT EXISTS date_of_birth DATE'))
                conn.execute(db.text('ALTER TABLE "user" ADD COLUMN IF NOT EXISTS phone VARCHAR(20)'))
                conn.execute(db.text('ALTER TABLE customer ADD COLUMN IF NOT EXISTS phone2 VARCHAR(20)'))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS lead_target INTEGER DEFAULT 0'))
                conn.execute(db.text('ALTER TABLE monthly_target ADD COLUMN IF NOT EXISTS conversion_target INTEGER DEFAULT 0'))
                conn.commit()
        except Exception as e:
            print(f'monthly_target table: {e}')
        # Create desk_note table if not exists
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS desk_note (
                        id SERIAL PRIMARY KEY,
                        user_id INTEGER REFERENCES "user"(id),
                        text TEXT NOT NULL,
                        reminder_date DATE,
                        mention_user_id INTEGER REFERENCES "user"(id),
                        is_done BOOLEAN DEFAULT FALSE,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                """))
                conn.commit()
        except Exception as e:
            print(f'desk_note table: {e}')
        
        # Add revenue column to job table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS revenue FLOAT DEFAULT 0'))
                conn.commit()
                print('✓ Revenue column migration completed')
        except Exception as e:
            print(f'Revenue column migration error: {e}')
        
        # Add partner commission columns to job table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_commission_expected BOOLEAN DEFAULT FALSE'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_name VARCHAR(100)'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_amount FLOAT'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_due_date DATE'))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_status VARCHAR(20) DEFAULT \'Pending\''))
                conn.execute(db.text('ALTER TABLE job ADD COLUMN IF NOT EXISTS partner_received_date DATE'))
                conn.commit()
                print('✓ Partner commission columns migration completed')
        except Exception as e:
            print(f'Partner commission migration error: {e}')
        
        # Create partner table
        try:
            with db.engine.connect() as conn:
                conn.execute(db.text("""
                    CREATE TABLE IF NOT EXISTS partner (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(100) NOT NULL UNIQUE,
                        active BOOLEAN DEFAULT TRUE,
                        created_at TIMESTAMP DEFAULT NOW()
                    )
                """))
                conn.commit()
                print('✓ Partner table created')
        except Exception as e:
            print(f'Partner table creation error: {e}')
        
        migrations = [
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS phone2 VARCHAR(20)',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS amount_invoiced FLOAT DEFAULT 0',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS amount_received FLOAT DEFAULT 0',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_approved_by INTEGER',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_approved_at TIMESTAMP',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS num_persons INTEGER DEFAULT 1',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS finance_notes TEXT',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS service_type VARCHAR(100)',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS due_date TIMESTAMP',
            'ALTER TABLE sub_task ADD COLUMN IF NOT EXISTS priority VARCHAR(20) DEFAULT \'Medium\'',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS final_remarks TEXT',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS future_work_notes TEXT',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS completed_at TIMESTAMP',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS service_note VARCHAR(200)',
            'ALTER TABLE job_update ADD COLUMN IF NOT EXISTS status_note VARCHAR(100)',
            'ALTER TABLE lead ADD COLUMN IF NOT EXISTS campaign VARCHAR(100)',
            """CREATE TABLE IF NOT EXISTS campaign (
                id SERIAL PRIMARY KEY,
                name VARCHAR(100) NOT NULL UNIQUE
            )""",

            'ALTER TABLE job ADD COLUMN IF NOT EXISTS status VARCHAR(50) DEFAULT \'Assigned\'',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS file_name VARCHAR(255)',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS file_url TEXT',
            'ALTER TABLE document ADD COLUMN IF NOT EXISTS cloudinary_public_id VARCHAR(255)',
            'ALTER TABLE job ADD COLUMN IF NOT EXISTS revenue_date DATE',
            'ALTER TABLE activity_log ADD COLUMN IF NOT EXISTS off_day VARCHAR(20)',
            'ALTER TABLE job_type ADD COLUMN IF NOT EXISTS default_days INTEGER DEFAULT 1',
            '''CREATE TABLE IF NOT EXISTS activity_type (
                id SERIAL PRIMARY KEY,
                field_key VARCHAR(50) UNIQUE NOT NULL,
                label VARCHAR(150) NOT NULL,
                weekly_target FLOAT DEFAULT 5,
                sort_order INTEGER DEFAULT 0,
                active BOOLEAN DEFAULT TRUE
            )''',
            'ALTER TABLE activity_type ADD COLUMN IF NOT EXISTS weekly_target FLOAT DEFAULT 5',
            'UPDATE \"user\" SET role = \'sales\' WHERE role = \'staff\'',

        ]
        for sql in migrations:
            try:
                with db.engine.connect() as conn:
                    conn.execute(db.text(sql))
                    conn.commit()
                    print(f'Migration OK: {sql[:60]}')
            except Exception as e:
                print(f'Migration skip: {sql[:60]} — {e}')
        try:
            admin = User.query.filter_by(email='admin@tahfeel.ae').first()
            if not admin:
                new_admin = User(
                    name='Admin-Tahfeel', email='admin@tahfeel.ae',
                    password=generate_password_hash('tahfeel2026'), role='admin'
                )
                db.session.add(new_admin)
                db.session.commit()
                print('Admin user created')
            elif admin.name == 'Admin':
                # Update existing admin name if it's still 'Admin'
                admin.name = 'Admin-Tahfeel'
                db.session.commit()
                print('Admin user name updated to Admin-Tahfeel')
            
            if Service.query.count() == 0:
                for s in ['Trade License', 'Family Visa', 'PRO Services', 'Healthcare License', 'Umrah Package', 'Other']:
                    db.session.add(Service(name=s))
                db.session.commit()
                print('Default services created')
            if Source.query.count() == 0:
                for s in ['Walk-in', 'WhatsApp', 'Referral', 'Social Media', 'Website', 'Other']:
                    db.session.add(Source(name=s))
                db.session.commit()
                print('Default sources created')
            if ServiceType.query.count() == 0:
                for jt in ['Trade License', 'Family Visa', 'PRO Services', 'Healthcare License', 'Umrah Package', 'Other']:
                    db.session.add(ServiceType(name=jt))
                db.session.commit()
                print('Default job types created')
            if ActivityType.query.count() == 0:
                for i, (key, label, target) in enumerate(ACTIVITY_DEFAULTS):
                    db.session.add(ActivityType(field_key=key, label=label, weekly_target=target, sort_order=i))
                db.session.commit()
                print('Default activity types seeded')
            if DocType.query.count() == 0:
                for dt in ['Trade License', 'Emirates ID', 'Passport', 'Visa', 'Medical Certificate', 'Insurance', 'Contract', 'NOC', 'Ejari', 'Other']:
                    db.session.add(DocType(name=dt))
                db.session.commit()
                print('Default doc types created')
        except Exception as e:
            db.session.rollback()
            print(f'Init db error: {e}')

@app.route('/admin/targets', methods=['GET','POST'])
@login_required
@admin_required
def set_targets():
    now = now_dubai()
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    if request.method == 'POST':
        month = int(request.form.get('month', now.month))
        year = int(request.form.get('year', now.year))
        # Get all active users (including admin)
        all_users = User.query.filter_by(active=True).all()
        saved = 0
        for u in all_users:
            val = request.form.get(f'amount_{u.id}', '').strip()
            amount_t = float(val) if val else 0.0
            try:
                t = MonthlyTarget.query.filter_by(user_id=u.id, month=month, year=year).first()
                if t:
                    t.amount_target = amount_t
                else:
                    t = MonthlyTarget(user_id=u.id, month=month, year=year,
                                     lead_target=0, conversion_target=0, amount_target=amount_t)
                    db.session.add(t)
                saved += 1
            except Exception as e:
                db.session.rollback()
                flash(f'Error saving target for {u.name}: {e}')
        try:
            db.session.commit()
            flash(f'Targets saved for {month}/{year}.')
        except Exception as e:
            db.session.rollback()
            flash(f'Save failed: {e}')
        return redirect(url_for('set_targets', month=month, year=year))
    users = User.query.filter_by(active=True).order_by(User.name).all()
    targets = {t.user_id: t for t in MonthlyTarget.query.filter_by(month=month, year=year).all()}
    return render_template('targets.html', users=users, targets=targets, month=month, year=year, now=now)

@app.route('/desk', methods=['GET','POST'])
@login_required
def my_desk():
    now = now_dubai()
    user_id = session['user_id']
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            text = request.form.get('text','').strip() or '📌'
            reminder_date = request.form.get('reminder_date','').strip()
            mention_user_ids = request.form.getlist('mention_user_ids')
            rd = datetime.strptime(reminder_date, '%Y-%m-%d').date() if reminder_date else None
            if mention_user_ids:
                # Create one note per mentioned user
                for mid in mention_user_ids:
                    note = DeskNote(user_id=user_id, text=text, reminder_date=rd, mention_user_id=int(mid))
                    db.session.add(note)
            else:
                note = DeskNote(user_id=user_id, text=text, reminder_date=rd, mention_user_id=None)
                db.session.add(note)
            db.session.commit()
        elif action == 'done':
            note_id = request.form.get('note_id')
            note = DeskNote.query.get(note_id)
            if note and (note.user_id == user_id or note.mention_user_id == user_id):
                note.is_done = not note.is_done
                db.session.commit()
        elif action == 'delete':
            note_id = request.form.get('note_id')
            note = DeskNote.query.get(note_id)
            if note and (note.user_id == user_id or note.mention_user_id == user_id):
                db.session.delete(note)
                db.session.commit()
        return redirect(url_for('my_desk'))

    # My notes + mentions
    my_notes = DeskNote.query.filter_by(user_id=user_id).order_by(DeskNote.is_done, DeskNote.reminder_date.asc().nullslast(), DeskNote.created_at.desc()).all()
    mentions = DeskNote.query.filter_by(mention_user_id=user_id, is_done=False).order_by(DeskNote.created_at.desc()).all()
    all_users = User.query.filter_by(active=True).filter(User.id != user_id).order_by(User.name).all()
    # Monthly targets + workload
    target = MonthlyTarget.query.filter_by(user_id=user_id, month=now.month, year=now.year).first()
    my_jobs_all = Job.query.filter_by(assigned_to=user_id).all()
    invoiced_actual = sum((j.amount_invoiced or 0) for j in my_jobs_all if j.status not in ['Pending Finance Approval'])
    closed_actual = sum((j.amount_received or 0) for j in my_jobs_all if j.status == 'Closed')
    amount_target = (target.amount_target or 0) if target else 0
    # Workload this month
    my_leads_month = Lead.query.filter_by(assigned_to=user_id).filter(
        db.extract('month', Lead.created_at) == now.month,
        db.extract('year', Lead.created_at) == now.year
    ).all()
    my_leads_count = len(my_leads_month)
    my_conversions_count = len([l for l in my_leads_month if l.status == 'Converted'])
    my_lost_count = len([l for l in my_leads_month if l.status == 'Lost'])
    my_overdue_leads = len([l for l in my_leads_month if l.due_date and l.due_date < now and l.status not in ['Converted','Lost']])
    my_active_tasks = len([j for j in my_jobs_all if j.status not in ['Done','Closed','Pending Finance Approval']])
    my_overdue_tasks = len([j for j in my_jobs_all if j.due_date and j.due_date < now and j.status not in ['Done','Closed','Pending Finance Approval']])

    # Create table if not exists
    try:
        db.session.execute(db.text('SELECT 1 FROM desk_note LIMIT 1'))
    except:
        db.session.rollback()
        with db.engine.connect() as conn:
            conn.execute(db.text("""
                CREATE TABLE IF NOT EXISTS desk_note (
                    id SERIAL PRIMARY KEY,
                    user_id INTEGER REFERENCES "user"(id),
                    text TEXT NOT NULL,
                    reminder_date DATE,
                    mention_user_id INTEGER REFERENCES "user"(id),
                    is_done BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMP DEFAULT NOW()
                )
            """))
            conn.commit()

    try:
        session['unread_mentions'] = len(mentions)
    except:
        pass
    return render_template('my_desk.html', my_notes=my_notes, mentions=mentions,
                           all_users=all_users, now=now,
                           invoiced_actual=invoiced_actual, closed_actual=closed_actual,
                           amount_target=amount_target,
                           my_leads_count=my_leads_count, my_conversions_count=my_conversions_count,
                           my_lost_count=my_lost_count, my_overdue_leads=my_overdue_leads,
                           my_active_tasks=my_active_tasks, my_overdue_tasks=my_overdue_tasks)

@app.route('/check-birthdays')
@login_required
def check_birthdays():
    today = now_dubai()
    try:
        result = db.session.execute(db.text("SELECT id, name, date_of_birth FROM customer WHERE date_of_birth IS NOT NULL")).fetchall()
        out = f"<b>Today (Dubai): {today.day}/{today.month}/{today.year}</b><br><br>Customers with DOB ({len(result)}):<br>"
        for r in result:
            dob = r[2]
            match = dob and dob.month == today.month and dob.day == today.day
            out += f"- {r[1]}: {dob} {'✅ BIRTHDAY TODAY' if match else ''}<br>"
        if not result:
            out += "No customers have DOB set yet."
        return out
    except Exception as e:
        return f'Error: {e}'
        
@app.route('/invoice')
@login_required
def invoice_generator():
    services = [s.name for s in Service.query.order_by(Service.name).all()]
    return render_template('invoice_generator.html', services=services)
   
@app.route('/admin/backup/export')
@login_required
@admin_required
def export_full_backup():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from flask import make_response
        import io

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="133E87", end_color="133E87", fill_type="solid")

        def style_headers(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Sheet 1: Leads
        try:
            ws1 = wb.active
            ws1.title = "Leads"
            users_map = {u.id: u.name for u in User.query.all()}
            style_headers(ws1, ['ID','Name','Company','Phone','Email','Service','Source','Campaign',
                                 'Status','Assigned To','Created Date','Due Date','Remarks'])
            for l in Lead.query.order_by(Lead.created_at.desc()).all():
                ws1.append([
                    l.id, l.name or '', l.company or '', l.phone or '', l.email or '',
                    l.service or '', l.source or '', l.campaign or '',
                    l.status or '', users_map.get(l.assigned_to, ''),
                    l.created_at.strftime('%d/%m/%Y %H:%M') if l.created_at else '',
                    l.due_date.strftime('%d/%m/%Y') if l.due_date else '',
                    l.remarks or ''
                ])
        except Exception as e:
            print(f"Error backing up Leads: {e}")
            flash(f'Warning: Leads backup incomplete - {str(e)}', 'warning')

        # Sheet 2: Customers
        try:
            ws2 = wb.create_sheet("Customers")
            style_headers(ws2, ['ID','Name','Company','Phone','Email','Source','Nationality',
                                 'Customer Type','Assigned To','Notes','Created Date'])
            for c in Customer.query.order_by(Customer.created_at.desc()).all():
                ws2.append([
                    c.id, c.name or '', c.company or '', c.phone or '', c.email or '',
                    c.source or '', c.nationality or '', c.customer_type or '',
                    users_map.get(c.assigned_to, ''), c.notes or '',
                    c.created_at.strftime('%d/%m/%Y %H:%M') if c.created_at else ''
                ])
        except Exception as e:
            print(f"Error backing up Customers: {e}")
            flash(f'Warning: Customers backup incomplete - {str(e)}', 'warning')

        # Sheet 3: Jobs
        try:
            ws3 = wb.create_sheet("Jobs")
            style_headers(ws3, ['ID','Customer','Company','Job Type','Assigned To','Created By',
                                 'Status','Invoiced (AED)','Received (AED)','Revenue (AED)',
                                 'Revenue Date','Partner Commission','Partner Received','Created Date','Due Date'])
            for j in Job.query.order_by(Job.created_at.desc()).all():
                try:
                    ws3.append([
                        j.id,
                        j.customer.name if j.customer else '',
                        j.customer.company if j.customer else '',
                        j.job_type or '',
                        users_map.get(j.assigned_to, ''),
                        users_map.get(j.created_by, ''),
                        j.status or '',
                        float(j.amount_invoiced or 0),
                        float(j.amount_received or 0),
                        float(j.revenue or 0),
                        j.revenue_date.strftime('%d/%m/%Y') if j.revenue_date else '',
                        'Yes' if j.has_partner_commission else 'No',
                        'Yes' if j.partner_commission_received else 'No',
                        j.created_at.strftime('%d/%m/%Y %H:%M') if j.created_at else '',
                        j.due_date.strftime('%d/%m/%Y') if j.due_date else ''
                    ])
                except Exception as row_error:
                    print(f"Error backing up Job ID {j.id}: {row_error}")
                    continue
        except Exception as e:
            print(f"Error backing up Jobs: {e}")
            flash(f'Warning: Jobs backup incomplete - {str(e)}', 'warning')

        # Sheet 4: Documents
        try:
            ws4 = wb.create_sheet("Documents")
            style_headers(ws4, ['ID','Doc Type','Owner Name','Belongs To','Customer',
                                 'Expiry Date','Notes','Added By','Created Date'])
            for d in Document.query.order_by(Document.created_at.desc()).all():
                try:
                    ws4.append([
                        d.id, d.doc_type or '', d.owner_name or '', d.belongs_to or '',
                        d.customer.name if d.customer else '',
                        d.expiry_date.strftime('%d/%m/%Y') if d.expiry_date else '',
                        d.notes or '', d.added_by or '',
                        d.created_at.strftime('%d/%m/%Y %H:%M') if d.created_at else ''
                    ])
                except Exception as row_error:
                    print(f"Error backing up Document ID {d.id}: {row_error}")
                    continue
        except Exception as e:
            print(f"Error backing up Documents: {e}")
            flash(f'Warning: Documents backup incomplete - {str(e)}', 'warning')

        # Sheet 5: Staff
        try:
            ws5 = wb.create_sheet("Staff")
            style_headers(ws5, ['ID','Name','Email','Role','Active'])
            for u in User.query.order_by(User.name).all():
                ws5.append([u.id, u.name, u.email, u.role, 'Yes' if u.active else 'No'])
        except Exception as e:
            print(f"Error backing up Staff: {e}")
            flash(f'Warning: Staff backup incomplete - {str(e)}', 'warning')

        # Mark backup date in session
        session['last_backup_date'] = now_dubai().strftime('%Y-%m-%d')
        session.modified = True

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"tahfeel_backup_{now_dubai().strftime('%Y%m%d_%H%M')}.xlsx"
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response
    
    except Exception as e:
        print(f"CRITICAL ERROR in backup export: {e}")
        import traceback
        traceback.print_exc()
        flash(f'Backup failed: {str(e)}. Please contact support.', 'error')
        return redirect(url_for('admin_panel'))


@app.route('/analytics')
@login_required
def analytics():
    if session.get('role') not in ['admin', 'finance']:
        flash('Access denied')
        return redirect(url_for('dashboard'))

    now = now_dubai()
    today = now.date()

    # Date range from request args — default this month
    period = request.args.get('period', 'this_month')
    if period == 'last_month':
        if now.month == 1:
            start = now.replace(year=now.year-1, month=12, day=1)
        else:
            start = now.replace(month=now.month-1, day=1)
        import calendar
        last_day = calendar.monthrange(start.year, start.month)[1]
        end = start.replace(day=last_day)
    elif period == 'this_year':
        start = now.replace(month=1, day=1)
        end = now
    else:  # this_month
        start = now.replace(day=1)
        end = now

    start_dt = start.replace(hour=0, minute=0, second=0) if hasattr(start, 'hour') else now.replace(day=1, hour=0, minute=0, second=0)
    end_dt = end.replace(hour=23, minute=59, second=59) if hasattr(end, 'hour') else now.replace(hour=23, minute=59, second=59)

    # All data for period
    all_leads = Lead.query.filter(Lead.created_at >= start_dt, Lead.created_at <= end_dt).all()
    all_jobs = Job.query.filter(Job.created_at >= start_dt, Job.created_at <= end_dt).all()
    all_users = User.query.filter_by(active=True).order_by(User.name).all()
    users_map = {u.id: u.name for u in User.query.all()}

    # ── Lead stats
    total_leads = len(all_leads)
    won_s = {'Won', 'Converted', 'Closed-Won'}
    lost_s = {'Lost', 'Rejected', 'Closed-Lost'}
    converted = [l for l in all_leads if l.status in won_s]
    lost = [l for l in all_leads if l.status in lost_s]
    conversion_rate = round(len(converted) / total_leads * 100, 1) if total_leads > 0 else 0

    # ── Revenue stats (match dashboard Finance card - count ALL jobs)
    total_invoiced = sum(j.amount_invoiced or 0 for j in all_jobs)
    total_received = sum(j.amount_received or 0 for j in all_jobs)
    total_outstanding = total_invoiced - total_received

    # ── Lead pipeline by status
    from collections import defaultdict, Counter
    pipeline = Counter(l.status for l in all_leads)

    # ── Top services
    service_counts = Counter(l.service for l in all_leads if l.service)
    top_services = service_counts.most_common(6)

    # ── Top sources
    source_counts = Counter(l.source for l in all_leads if l.source)
    top_sources = source_counts.most_common(6)

    # ── Campaign performance
    campaign_counts = Counter(l.campaign for l in all_leads if l.campaign)
    top_campaigns = campaign_counts.most_common(5)

    # ── Monthly revenue trend (last 6 months)
    monthly_revenue = []
    for i in range(5, -1, -1):
        if now.month - i <= 0:
            m = now.month - i + 12
            y = now.year - 1
        else:
            m = now.month - i
            y = now.year
        month_jobs = Job.query.filter(
            db.extract('month', Job.created_at) == m,
            db.extract('year', Job.created_at) == y,
            Job.status.notin_(['Pending Finance Approval'])
        ).all()
        inv = sum(j.amount_invoiced or 0 for j in month_jobs)
        rec = sum(j.amount_received or 0 for j in month_jobs)
        import calendar
        monthly_revenue.append({
            'month': calendar.month_abbr[m],
            'invoiced': inv,
            'received': rec,
        })

    # ── Staff performance
    staff_stats = []
    # Get monthly targets
    month = int(request.args.get('month', now.month))
    year = int(request.args.get('year', now.year))
    targets = {t.user_id: t.amount_target or 0 for t in MonthlyTarget.query.filter_by(month=month, year=year).all()}
    
    for u in all_users:
        if u.role != 'sales':  # Only Sales role
            continue
        u_leads = [l for l in all_leads if l.assigned_to == u.id]
        u_sales = [j for j in all_jobs if j.customer and j.customer.assigned_to == u.id]
        u_inv = sum(j.amount_invoiced or 0 for j in u_sales if j.status not in ['Pending Finance Approval'])
        u_conv = len([l for l in u_leads if l.status in won_s])
        conv_rate = round(u_conv / len(u_leads) * 100) if u_leads else 0
        
        # Count pending leads (not contacted yet - status is "New")
        u_pending = len([l for l in u_leads if l.status == 'New'])
        
        # Calculate revenue from closed jobs + partial revenues
        try:
            u_revenue = sum(j.revenue or 0 for j in u_sales if j.status == 'Closed')
            # Add partial revenues from non-closed jobs
            for j in u_sales:
                if j.status != 'Closed':
                    u_revenue += sum(pr.amount for pr in j.partial_revenues)
        except:
            u_revenue = 0
        
        staff_stats.append({
            'name': u.name,
            'role': u.role,
            'leads': len(u_leads),
            'converted': u_conv,
            'conv_rate': conv_rate,
            'invoiced': u_inv,
            'revenue': u_revenue,
            'target': targets.get(u.id, 0),
            'pending': u_pending,
        })
    staff_stats.sort(key=lambda x: x['revenue'], reverse=True)

    # Max values for bar scaling
    max_service = top_services[0][1] if top_services else 1
    max_source = top_sources[0][1] if top_sources else 1
    max_pipeline = max(pipeline.values()) if pipeline else 1
    max_rev = max((m['invoiced'] for m in monthly_revenue), default=1) or 1

    # ── Operations stats
    total_jobs = len(all_jobs)
    completed_jobs = [j for j in all_jobs if j.status in ['Done','Closed']]
    active_jobs_ops = [j for j in all_jobs if j.status not in ['Done','Closed','Pending Finance Approval']]
    overdue_jobs = [j for j in all_jobs if j.due_date and j.due_date.date() < today and j.status not in ['Done','Closed']]
    job_type_counts = Counter(j.job_type for j in all_jobs if j.job_type)
    top_job_types = job_type_counts.most_common(6)
    max_job_type = top_job_types[0][1] if top_job_types else 1
    job_status_counts = Counter(j.status for j in all_jobs)

    # Avg days to complete
    completion_days = []
    for j in completed_jobs:
        if j.created_at and j.completed_at:
            completion_days.append((j.completed_at - j.created_at).days)
    avg_completion = round(sum(completion_days) / len(completion_days), 1) if completion_days else 0

    # Avg completion days per job type
    from collections import defaultdict
    jtype_days = defaultdict(list)
    for j in Job.query.filter(Job.status.in_(['Done','Closed']), Job.completed_at.isnot(None)).all():
        if j.created_at and j.completed_at and j.job_type:
            jtype_days[j.job_type].append((j.completed_at - j.created_at).days)
    avg_by_job_type = sorted(
        [{'type': jt, 'avg': round(sum(days)/len(days), 1), 'count': len(days)}
         for jt, days in jtype_days.items()],
        key=lambda x: x['avg'], reverse=True
    )

    # ── Documents stats
    from datetime import timedelta as _td
    all_docs = Document.query.all()
    total_docs = len(all_docs)
    expired_docs = [d for d in all_docs if d.expiry_date and d.expiry_date.date() < today]
    expiring_30 = [d for d in all_docs if d.expiry_date and 0 <= (d.expiry_date.date() - today).days <= 30]
    expiring_60 = [d for d in all_docs if d.expiry_date and 31 <= (d.expiry_date.date() - today).days <= 60]
    expiring_90 = [d for d in all_docs if d.expiry_date and 61 <= (d.expiry_date.date() - today).days <= 90]
    doc_type_counts = Counter(d.doc_type for d in all_docs if d.doc_type)
    top_doc_types = doc_type_counts.most_common(6)

    # Tab from request
    tab = request.args.get('tab', 'overview')
    
    # ── Lead breakdown by staff and status (for pivot table)
    # Get all statuses and staff
    all_statuses = sorted(set(l.status for l in all_leads if l.status), key=lambda s: (s != 'New', s))  # "New" first, then alphabetical
    sales_staff = [u for u in all_users if u.role == 'sales']  # Only Sales role, not admin
    
    # Create breakdown: {status: {staff_name: count}}
    lead_breakdown = {}
    for status in all_statuses:
        lead_breakdown[status] = {}
        for staff in sales_staff:
            count = len([l for l in all_leads if l.status == status and l.assigned_to == staff.id])
            lead_breakdown[status][staff.name] = count
    
    # Calculate totals
    status_totals = {status: sum(lead_breakdown[status].values()) for status in all_statuses}
    staff_totals = {staff.name: sum(lead_breakdown[status].get(staff.name, 0) for status in all_statuses) for staff in sales_staff}
    grand_total = sum(status_totals.values())

    return render_template('analytics.html',
        now=now, period=period, tab=tab,
        total_leads=total_leads, converted=len(converted), lost=len(lost),
        conversion_rate=conversion_rate,
        total_invoiced=total_invoiced, total_received=total_received,
        total_outstanding=total_outstanding,
        pipeline=pipeline, top_services=top_services, top_sources=top_sources,
        top_campaigns=top_campaigns, monthly_revenue=monthly_revenue,
        staff_stats=staff_stats,
        max_service=max_service, max_source=max_source,
        max_pipeline=max_pipeline, max_rev=max_rev,
        users_map=users_map,
        lead_breakdown=lead_breakdown, all_statuses=all_statuses, sales_staff=sales_staff,
        status_totals=status_totals, staff_totals=staff_totals, grand_total=grand_total,
        total_jobs=total_jobs, completed_jobs=len(completed_jobs),
        active_jobs_ops=len(active_jobs_ops), overdue_jobs=len(overdue_jobs),
        avg_completion=avg_completion, top_job_types=top_job_types, avg_by_job_type=avg_by_job_type,
        max_job_type=max_job_type, job_status_counts=job_status_counts,
        total_docs=total_docs, expired_docs=expired_docs,
        expiring_30=expiring_30, expiring_60=expiring_60, expiring_90=expiring_90,
        top_doc_types=top_doc_types,
    )

from reports import reports_bp
app.register_blueprint(reports_bp)

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
else:
    init_db()

@app.route('/partners')
@login_required
def partners():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    
    partners = Partner.query.order_by(Partner.name).all()
    return render_template('partners.html', partners=partners)

@app.route('/partners/add', methods=['POST'])
@login_required
def add_partner():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    name = request.form.get('name', '').strip()
    if not name:
        flash('Partner name is required.', 'error')
        return redirect(url_for('partners'))
    
    existing = Partner.query.filter_by(name=name).first()
    if existing:
        flash('Partner already exists.', 'error')
        return redirect(url_for('partners'))
    
    partner = Partner(name=name)
    db.session.add(partner)
    db.session.commit()
    flash(f'Partner "{name}" added successfully.')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/edit', methods=['POST'])
@login_required
def edit_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    new_name = request.form.get('name', '').strip()
    
    if not new_name:
        flash('Partner name cannot be empty.', 'error')
        return redirect(url_for('partners'))
    
    # Check if new name already exists (excluding current partner)
    existing = Partner.query.filter(Partner.name == new_name, Partner.id != partner_id).first()
    if existing:
        flash('Partner name already exists.', 'error')
        return redirect(url_for('partners'))
    
    old_name = partner.name
    partner.name = new_name
    db.session.commit()
    flash(f'Partner renamed from "{old_name}" to "{new_name}".')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/delete', methods=['POST'])
@login_required
def delete_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    
    # Check if partner has any jobs associated
    jobs_with_partner = Job.query.filter_by(partner_name=partner.name).first()
    if jobs_with_partner:
        flash(f'Cannot delete "{partner.name}" - has associated transactions. Deactivate instead.', 'error')
        return redirect(url_for('partners'))
    
    name = partner.name
    db.session.delete(partner)
    db.session.commit()
    flash(f'Partner "{name}" deleted successfully.')
    return redirect(url_for('partners'))

@app.route('/partners/<int:partner_id>/toggle', methods=['POST'])
@login_required
def toggle_partner(partner_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partners'))
    
    partner = Partner.query.get_or_404(partner_id)
    partner.active = not partner.active
    db.session.commit()
    status = 'activated' if partner.active else 'deactivated'
    flash(f'Partner "{partner.name}" {status}.')
    return redirect(url_for('partners'))

@app.route('/partner-commissions')
@login_required
def partner_commissions():
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('dashboard'))
    
    try:
        # Get all jobs with pending partner commissions
        all_pending = Job.query.filter_by(partner_commission_expected=True, partner_status='Pending').all()
    except Exception as e:
        print(f"Error querying pending jobs: {e}")
        all_pending = []
    
    # Filter options
    partner_filter = request.args.get('partner', '')
    status_filter = request.args.get('status', 'pending')
    
    # Apply filters
    try:
        if status_filter == 'pending':
            jobs = all_pending
        elif status_filter == 'received':
            jobs = Job.query.filter_by(partner_commission_expected=True, partner_status='Received').all()
        else:  # all
            jobs = Job.query.filter_by(partner_commission_expected=True).all()
    except Exception as e:
        print(f"Error filtering jobs: {e}")
        jobs = []
    
    if partner_filter:
        jobs = [j for j in jobs if j.partner_name == partner_filter]
    
    # Get unique partners for filter dropdown
    try:
        all_partners = Partner.query.filter_by(active=True).order_by(Partner.name).all()
    except Exception as e:
        print(f"Error querying partners: {e}")
        all_partners = []
    
    # Calculate totals
    try:
        total_pending = sum((j.partner_amount or 0) for j in all_pending)
        total_received = sum((j.partner_amount or 0) for j in Job.query.filter_by(partner_commission_expected=True, partner_status='Received').all())
    except Exception as e:
        print(f"Error calculating totals: {e}")
        total_pending = 0
        total_received = 0
    
    now = now_dubai()
    
    return render_template('partner_commissions.html', 
                          jobs=jobs, 
                          all_partners=all_partners,
                          partner_filter=partner_filter,
                          status_filter=status_filter,
                          total_pending=total_pending,
                          total_received=total_received,
                          now=now)

@app.route('/partner-commissions/<int:job_id>/mark-received', methods=['POST'])
@login_required
def mark_partner_received(job_id):
    if session['role'] not in ['admin', 'finance']:
        flash('Access denied.')
        return redirect(url_for('partner_commissions'))
    
    job = Job.query.get_or_404(job_id)
    
    if not job.partner_commission_expected or job.partner_status != 'Pending':
        flash('This task does not have a pending partner commission.', 'error')
        return redirect(url_for('partner_commissions'))
    
    # Mark as received
    job.partner_status = 'Received'
    job.partner_received_date = now_dubai().date()
    job.revenue = job.partner_amount  # NOW count the revenue!
    job.revenue_date = now_dubai().date()  # Revenue counted TODAY (cash-basis)
    job.status = 'Closed'  # Remove "Pending Partner Commission" from status
    
    # Add timeline update
    remark = f'Partner commission RECEIVED from {job.partner_name}: AED {job.partner_amount:,.0f}. Revenue now counted for {now_dubai().strftime("%B %Y")} (cash-basis). Marked by {session["user_name"]}.'
    update = JobUpdate(job_id=job.id, status='Closed', remark=remark, staff_name=session['user_name'])
    db.session.add(update)
    db.session.commit()
    
    flash(f'Partner commission of AED {job.partner_amount:,.0f} from {job.partner_name} marked as received. Revenue added to totals.')
    return redirect(url_for('partner_commissions'))

# ── Admin Panel Partner Routes (simpler pattern)
@app.route('/admin/partner/add', methods=['POST'])
@login_required
@admin_required
def admin_add_partner():
    name = request.form.get('name', '').strip()
    if name:
        existing = Partner.query.filter_by(name=name).first()
        if not existing:
            partner = Partner(name=name)
            db.session.add(partner)
            db.session.commit()
            flash(f'Partner "{name}" added.')
        else:
            flash('Partner already exists.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/partner/<int:partner_id>/edit', methods=['POST'])
@login_required
@admin_required
def admin_edit_partner(partner_id):
    partner = Partner.query.get_or_404(partner_id)
    new_name = request.form.get('name', '').strip()
    if new_name:
        existing = Partner.query.filter(Partner.name == new_name, Partner.id != partner_id).first()
        if not existing:
            partner.name = new_name
            db.session.commit()
            flash(f'Partner updated.')
        else:
            flash('Partner name already exists.', 'error')
    return redirect(url_for('admin_panel'))

@app.route('/admin/partner/<int:partner_id>/delete')
@login_required
@admin_required
def admin_delete_partner(partner_id):
    partner = Partner.query.get_or_404(partner_id)
    jobs_with_partner = Job.query.filter_by(partner_name=partner.name).first()
    if not jobs_with_partner:
        db.session.delete(partner)
        db.session.commit()
        flash(f'Partner "{partner.name}" deleted.')
    else:
        flash('Cannot delete - has associated tasks.', 'error')
    return redirect(url_for('admin_panel'))

# ══════════════════════════════════════════════════════════════════════════════
# TEMPORARY ADMIN ROUTE - Fix April 30 Revenue Dates
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
